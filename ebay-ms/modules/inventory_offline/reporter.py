"""
modules/inventory_offline/reporter.py

Day 23: 库存报表

功能：
- 当前库存快照（所有 SKU + 可用数量 + 位置分布 + 进货价 + 库存金额）
- 出入库明细（日期范围筛选）
- 库存变动趋势（按 SKU 过去 N 天）
- Excel 导出
"""

from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

from loguru import logger as log

# ── 数据模型 ────────────────────────────────────────────────────────────────

@dataclass
class StockSnapshotItem:
    """当前库存快照的一行。"""
    sku: str
    title: str | None
    available_quantity: int
    locations: dict[str, int]         # location -> quantity
    cost_price: Decimal
    inventory_value: Decimal         # available_quantity * cost_price
    last_inbound_at: datetime | None
    last_outbound_at: datetime | None


@dataclass
class MovementItem:
    """出入库明细的一行。"""
    occurred_at: datetime
    movement_type: str               # IN / OUT / ADJUST / RETURN
    sku: str
    quantity: int                    # 变动数量（ADJUST 可为负）
    related_order: str | None
    location: str | None
    operator: str | None
    note: str | None
    unit_cost: Decimal | None        # 入库时单价（IN/RETURN 类型记录）
    total_cost: Decimal | None       # quantity * unit_cost


@dataclass
class TrendItem:
    """库存变动趋势的一行。"""
    date: str                        # YYYY-MM-DD
    sku: str
    opening: int                     # 当日开盘库存
    inbound: int                     # 当日入库合计
    outbound: int                    # 当日出库合计（正数）
    adjustment: int                 # 当日调整合计
    closing: int                     # 当日收盘库存


# ── Reporter ────────────────────────────────────────────────────────────────

class InventoryReporter:
    """库存报表生成器。"""

    def __init__(self):
        from modules.inventory_offline import InboundService
        self._svc = InboundService()

    # ── 当前库存快照 ───────────────────────────────────────────────

    def get_stock_snapshot(self) -> list[StockSnapshotItem]:
        """
        返回所有 SKU 的当前库存快照。
        """
        all_stock = self._svc.get_all_stock()
        items: list[StockSnapshotItem] = []

        for entry in all_stock:
            sku = entry["sku"]
            available = entry["available_quantity"]
            cost_price = entry.get("cost_price") or Decimal("0")

            items.append(StockSnapshotItem(
                sku=sku,
                title=entry.get("title"),
                available_quantity=available,
                locations=entry.get("locations") or {},
                cost_price=cost_price,
                inventory_value=cost_price * available,
                last_inbound_at=entry.get("last_inbound_at"),
                last_outbound_at=entry.get("last_outbound_at"),
            ))

        # 按库存金额降序
        items.sort(key=lambda x: x.inventory_value, reverse=True)
        return items

    # ── 出入库明细 ─────────────────────────────────────────────────

    def get_movements(
        self,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        sku: str | None = None,
        movement_type: str | None = None,
        related_order: str | None = None,
        limit: int = 1000,
    ) -> list[MovementItem]:
        """
        查询出入库明细记录。

        Args:
            start_date: 起始时间（含）
            end_date: 结束时间（含）
            sku: 按 SKU 筛选
            movement_type: 按类型筛选（IN/OUT/ADJUST/RETURN）
            related_order: 按订单号筛选
            limit: 最大返回条数

        Returns:
            list[MovementItem]，按时间倒序
        """
        from core.database.connection import get_session
        from core.models import Inventory

        with get_session() as sess:
            query = sess.query(Inventory).order_by(Inventory.occurred_at.desc())

            if start_date:
                query = query.filter(Inventory.occurred_at >= start_date)
            if end_date:
                query = query.filter(Inventory.occurred_at <= end_date)
            if sku:
                query = query.filter(Inventory.sku == sku)
            if movement_type:
                query = query.filter(Inventory.type == movement_type)
            if related_order:
                query = query.filter(Inventory.related_order == related_order)

            rows = query.limit(limit).all()

            # 在 session 关闭前构建 MovementItem，避免 DetachedInstanceError
            result = []
            for r in rows:
                result.append(MovementItem(
                    occurred_at=r.occurred_at,
                    movement_type=r.type.value,
                    sku=r.sku,
                    quantity=r.quantity,
                    related_order=r.related_order,
                    location=r.location,
                    operator=r.operator,
                    note=r.note,
                    unit_cost=None,
                    total_cost=None,
                ))
            return result

    # ── 库存变动趋势 ────────────────────────────────────────────────

    def get_trend(self, sku: str, lookback_days: int = 30) -> list[TrendItem]:
        """
        生成指定 SKU 过去 N 天的库存变动趋势（每日汇总）。
        """
        from core.database.connection import get_session
        from core.models import Inventory

        # 使用 naive datetime 避免 SQLite 时区比较问题
        now_naive = datetime.now()
        start_date = now_naive - timedelta(days=lookback_days)
        end_date = now_naive

        with get_session() as sess:
            records = (
                sess.query(Inventory)
                .filter(
                    Inventory.sku == sku,
                    Inventory.occurred_at >= start_date,
                    Inventory.occurred_at <= end_date,
                )
                .order_by(Inventory.occurred_at.desc())
                .all()
            )

            if not records:
                return []

            # 按日期分组（在 session 关闭前完成所有属性访问）
            from collections import defaultdict
            daily: dict[str, dict] = defaultdict(
                lambda: {"in": 0, "out": 0, "adjust": 0, "return": 0}
            )

            for r in records:
                date_key = r.occurred_at.strftime("%Y-%m-%d")
                qty = r.quantity
                t = r.type.value
                if t == "in":
                    daily[date_key]["in"] += qty
                elif t == "out":
                    daily[date_key]["out"] += abs(qty)
                elif t == "adjust":
                    daily[date_key]["adjust"] += qty
                elif t == "return":
                    daily[date_key]["return"] += qty

            sorted_dates = sorted(daily.keys(), reverse=True)

        stock_info = self._svc.get_stock(sku)
        current_available = stock_info["available_quantity"]

        trend: list[TrendItem] = []
        closing = current_available

        for date_key in sorted_dates:
            d = daily[date_key]
            inbound = d["in"]
            outbound = d["out"]
            adjustment = d["adjust"]
            opening = closing - inbound - d["return"] + outbound - adjustment

            trend.append(TrendItem(
                date=date_key,
                sku=sku,
                opening=opening,
                inbound=inbound,
                outbound=outbound,
                adjustment=adjustment,
                closing=closing,
            ))
            closing = opening

        trend.reverse()
        return trend

    # ── Excel 导出 ───────────────────────────────────────────────────

    def export_snapshot_to_excel(self, path: str | Path) -> None:
        """导出当前库存快照为 Excel。"""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "库存快照"

        items = self.get_stock_snapshot()

        headers = ["SKU", "商品名称", "可用数量", "位置分布", "进货价", "库存金额 (JPY)", "最后入库", "最后出库"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item.sku)
            ws.cell(row=row_idx, column=2, value=item.title or "")
            ws.cell(row=row_idx, column=3, value=item.available_quantity)
            # 位置分布：格式化为 "A-1:5, B-2:3"
            loc_str = ", ".join(f"{k}:{v}" for k, v in sorted(item.locations.items()))
            ws.cell(row=row_idx, column=4, value=loc_str)
            ws.cell(row=row_idx, column=5, value=float(item.cost_price) if item.cost_price else None)
            ws.cell(row=row_idx, column=6, value=float(item.inventory_value))
            ws.cell(row=row_idx, column=7, value=str(item.last_inbound_at or ""))
            ws.cell(row=row_idx, column=8, value=str(item.last_outbound_at or ""))

        total_value = sum(i.inventory_value for i in items)
        total_qty = sum(i.available_quantity for i in items)
        summary_row = len(items) + 2
        ws.cell(row=summary_row, column=1, value="合计")
        ws.cell(row=summary_row, column=3, value=total_qty)
        ws.cell(row=summary_row, column=4, value=float(total_value))
        ws.cell(row=summary_row, column=3).font = Font(bold=True)
        ws.cell(row=summary_row, column=4).font = Font(bold=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20

        wb.save(path)
        log.info(f"库存快照已导出：{path}")

    def export_movements_to_excel(
        self,
        path: str | Path,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        sku: str | None = None,
    ) -> None:
        """导出出入库明细为 Excel。"""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "出入库明细"

        items = self.get_movements(
            start_date=start_date,
            end_date=end_date,
            sku=sku,
        )

        headers = ["时间", "类型", "SKU", "数量", "单件成本", "总成本", "订单号", "操作人", "备注"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        type_colors = {
            "in": "C6EFCE",    # 绿
            "out": "FFCCCC",   # 红
            "adjust": "FFEB9C",  # 黄
            "return": "DDEBF7",  # 蓝
        }

        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item.occurred_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row_idx, column=2, value=item.movement_type)
            ws.cell(row=row_idx, column=3, value=item.sku)
            ws.cell(row=row_idx, column=4, value=item.quantity)
            ws.cell(row=row_idx, column=5, value=float(item.unit_cost) if item.unit_cost else None)
            ws.cell(row=row_idx, column=6, value=float(item.total_cost) if item.total_cost else None)
            ws.cell(row=row_idx, column=7, value=item.related_order or "")
            ws.cell(row=row_idx, column=8, value=item.operator or "")
            ws.cell(row=row_idx, column=9, value=item.note or "")

            fill = PatternFill("solid", fgColor=type_colors.get(item.movement_type, "FFFFFF"))
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 20

        wb.save(path)
        log.info(f"出入库明细已导出：{path}（{len(items)} 条）")
