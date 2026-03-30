"""
tests/test_generator_spreadsheet.py — Excel 报表生成测试
"""
import os
import pytest
from unittest.mock import patch
from openpyxl import load_workbook
from db.db import init_db, insert, fetch_all
from generator.spreadsheet import generate_report


@pytest.fixture(scope='function')
def test_db(tmp_path, monkeypatch):
    """为每个测试创建独立的临时数据库"""
    db_path = tmp_path / "test_orders.db"

    from db import db as db_module

    def mock_get_db_path():
        return str(db_path)

    monkeypatch.setattr(db_module, 'get_db_path', mock_get_db_path)

    init_db()
    yield str(db_path)

    if db_path.exists():
        db_path.unlink()


@pytest.fixture
def sample_data(test_db):
    """创建示例数据"""
    # 插入 2024 年的 eBay 订单
    insert('ebay_orders', {
        'order_id': 'ORDER-001',
        'sale_date': '2024-01-15',
        'buyer_username': 'buyer1',
        'item_title': 'Test Item 1',
        'quantity': 1,
        'sale_price_usd': 100.0,
        'shipping_charged_usd': 10.0,
        'ebay_fee_usd': 10.0,
        'ebay_ad_fee_usd': 5.0,
    })

    insert('ebay_orders', {
        'order_id': 'ORDER-002',
        'sale_date': '2024-02-20',
        'buyer_username': 'buyer2',
        'item_title': 'Test Item 2',
        'quantity': 2,
        'sale_price_usd': 200.0,
        'shipping_charged_usd': 15.0,
        'ebay_fee_usd': 20.0,
        'ebay_ad_fee_usd': 10.0,
    })

    # 插入 2023 年的订单（不应该出现在 2024 年报中）
    insert('ebay_orders', {
        'order_id': 'ORDER-OLD',
        'sale_date': '2023-12-01',
        'buyer_username': 'buyer_old',
        'item_title': 'Old Item',
        'quantity': 1,
        'sale_price_usd': 50.0,
        'shipping_charged_usd': 5.0,
        'ebay_fee_usd': 5.0,
        'ebay_ad_fee_usd': 0.0,
    })

    # 插入 shipment（关联 ORDER-001）
    insert('shipments', {
        'id': 'cpass_TXN-001',
        'carrier': 'cpass_speedpak',
        'tracking_number': 'TRACK-001',
        'ebay_order_id': 'ORDER-001',
        'ship_date': '2024-01-16',
        'shipping_fee_usd': 8.0,
        'cpass_transaction_id': 'TXN-001',
    })

    # 插入未匹配的 shipment
    insert('shipments', {
        'id': 'cpass_TXN-UNMATCHED',
        'carrier': 'cpass_fedex',
        'tracking_number': 'TRACK-UNMATCHED',
        'ebay_order_id': None,  # 未匹配
        'ship_date': '2024-03-01',
        'shipping_fee_usd': 20.0,
        'cpass_transaction_id': 'TXN-UNMATCHED',
    })

    # 插入采购记录
    insert('purchases', {
        'id': 'amazon_JPN-001',
        'platform': 'amazon_jp',
        'purchase_date': '2024-01-10',
        'item_name': 'テスト商品 1',
        'quantity': 1,
        'total_price_jpy': 5000.0,
        'tax_jpy': 500.0,
        'order_number': 'JPN-001',
    })

    # 插入匹配关系
    insert('purchase_order_links', {
        'purchase_id': 'amazon_JPN-001',
        'ebay_order_id': 'ORDER-001',
        'match_method': 'sku',
        'confidence': 0.95,
    })

    return test_db


class TestGenerateReport:
    """Excel 报表生成测试"""

    def test_generate_report_basic(self, test_db, sample_data, tmp_path):
        """测试基本报表生成功能"""
        output_path = tmp_path / "tax_report_2024.xlsx"

        result_path = generate_report(2024, str(output_path))

        assert result_path == str(output_path)
        assert os.path.exists(result_path)

        # 验证生成的 Excel 文件
        wb = load_workbook(result_path)

        # 检查 Sheet1
        assert "订单明细" in wb.sheetnames
        ws1 = wb["订单明细"]

        # 验证表头
        headers = [cell.value for cell in ws1[1]]
        expected_headers = [
            "eBay 订单号", "成交日期", "商品名称", "数量", "销售价 (USD)", "买家运费 (USD)",
            "eBay 平台费 (USD)", "eBay 广告费 (USD)", "快递方式", "快递单号",
            "国际快递费 (USD)", "采购平台", "采购价 (JPY)", "采购消费税 (JPY)", "采购日期",
            "采购订单号", "汇率 (JPY/USD)", "净利润估算 (USD)", "匹配状态"
        ]
        assert headers == expected_headers

        # 验证表头样式（加粗、背景色、白色字体）
        for cell in ws1[1]:
            assert cell.font.bold is True
            assert cell.fill.start_color.rgb == "002F5496" or cell.fill.start_color.rgb == "2F5496"

        # 验证数据行数（2 条 2024 年订单）
        rows = list(ws1.iter_rows(values_only=True))
        assert len(rows) == 3  # 1 行表头 + 2 行数据

        # 验证第一条订单数据
        row1 = rows[1]
        assert row1[0] == 'ORDER-001'  # eBay 订单号
        assert row1[1] == '2024-01-15'  # 成交日期
        assert row1[2] == 'Test Item 1'  # 商品名称
        assert row1[3] == 1  # 数量
        assert row1[4] == 100.0  # 销售价
        assert row1[5] == 10.0  # 买家运费

        # 检查 Sheet2
        assert "未匹配记录" in wb.sheetnames
        ws2 = wb["未匹配记录"]

        # 验证未匹配记录
        unmatched_rows = list(ws2.iter_rows(values_only=True))
        assert len(unmatched_rows) == 2  # 1 行表头 + 1 条未匹配记录

    @patch('generator.spreadsheet.batch_get_rates')
    def test_generate_report_net_profit_calculation(self, mock_rates, test_db, sample_data, tmp_path):
        """测试净利润计算"""
        # Mock 汇率为固定 1/150
        mock_rates.return_value = {
            '2024-01-15': 1/150.0,
            '2024-02-20': 1/150.0,
        }

        output_path = tmp_path / "tax_report_2024.xlsx"
        generate_report(2024, str(output_path))

        wb = load_workbook(output_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # ORDER-001 的净利润计算：
        # sale_price=100 + shipping=10 - ebay_fee=10 - ebay_ad_fee=5 - shipping_fee=8 - purchase_jpy=5000*(1/150)=33.33
        # = 100 + 10 - 10 - 5 - 8 - 33.33 = 53.67
        row1 = rows[1]
        net_profit = row1[17]  # 净利润估算 (USD) 列（索引 17，因为汇率列在 16）
        assert abs(net_profit - 53.67) < 0.1

        # ORDER-002 的净利润计算（无采购）：
        # 200 + 15 - 20 - 10 - 0 - 0 = 185
        row2 = rows[2]
        net_profit2 = row2[17]
        assert net_profit2 == 185.0

    def test_generate_report_match_status(self, test_db, sample_data, tmp_path):
        """测试匹配状态"""
        output_path = tmp_path / "tax_report_2024.xlsx"
        generate_report(2024, str(output_path))

        wb = load_workbook(output_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # ORDER-001: 有采购有 shipment -> 已匹配
        row1 = rows[1]
        assert row1[18] == "已匹配"  # 匹配状态（索引 18）

        # ORDER-002: 无采购无 shipment -> 未匹配
        row2 = rows[2]
        assert row2[18] == "未匹配"

    def test_generate_report_excludes_other_year(self, test_db, sample_data, tmp_path):
        """测试报表不包含其他年份的数据"""
        output_path = tmp_path / "tax_report_2024.xlsx"
        generate_report(2024, str(output_path))

        wb = load_workbook(output_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # 验证不包含 2023 年的订单
        order_ids = [row[0] for row in rows[1:]]
        assert 'ORDER-OLD' not in order_ids

    def test_generate_report_unmatched_sheet(self, test_db, sample_data, tmp_path):
        """测试未匹配记录工作表"""
        output_path = tmp_path / "tax_report_2024.xlsx"
        generate_report(2024, str(output_path))

        wb = load_workbook(output_path)
        ws2 = wb["未匹配记录"]
        rows = list(ws2.iter_rows(values_only=True))

        # 验证表头
        assert rows[0][3] == 'ebay_order_id'

        # 验证未匹配记录内容
        assert len(rows) == 2  # 1 行表头 + 1 条未匹配记录
        assert rows[1][3] is None or rows[1][3] == ''  # ebay_order_id 为空

    def test_generate_report_empty_db(self, test_db, tmp_path):
        """测试空数据库生成报表"""
        output_path = tmp_path / "tax_report_empty.xlsx"
        result_path = generate_report(2024, str(output_path))

        assert os.path.exists(result_path)

        wb = load_workbook(result_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # 只有表头
        assert len(rows) == 1


class TestGenerateReportMonthFilter:
    """月份过滤功能测试"""

    def test_generate_month_filter(self, test_db, tmp_path):
        """月份过滤后只包含该月订单"""
        # 插入 1 月和 2 月订单各一条
        insert('ebay_orders', {
            'order_id': 'ORDER-JAN',
            'sale_date': '2024-01-15',
            'buyer_username': 'buyer_jan',
            'item_title': 'January Item',
            'quantity': 1,
            'sale_price_usd': 100.0,
            'shipping_charged_usd': 10.0,
            'ebay_fee_usd': 10.0,
            'ebay_ad_fee_usd': 5.0,
        })

        insert('ebay_orders', {
            'order_id': 'ORDER-FEB',
            'sale_date': '2024-02-20',
            'buyer_username': 'buyer_feb',
            'item_title': 'February Item',
            'quantity': 2,
            'sale_price_usd': 200.0,
            'shipping_charged_usd': 15.0,
            'ebay_fee_usd': 20.0,
            'ebay_ad_fee_usd': 10.0,
        })

        # 生成 2 月份报表
        output_path = tmp_path / "tax_report_2024-02.xlsx"
        generate_report(2024, str(output_path), month=2)

        wb = load_workbook(output_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # 验证只有 2 月订单（1 行表头 + 1 行数据）
        assert len(rows) == 2
        assert rows[1][0] == 'ORDER-FEB'  # eBay 订单号
        assert rows[1][1] == '2024-02-20'  # 成交日期

    def test_generate_month_no_data(self, test_db, tmp_path, capsys):
        """无数据时生成空报表不报错"""
        # 只插入 1 月订单
        insert('ebay_orders', {
            'order_id': 'ORDER-JAN',
            'sale_date': '2024-01-15',
            'buyer_username': 'buyer_jan',
            'item_title': 'January Item',
            'quantity': 1,
            'sale_price_usd': 100.0,
            'shipping_charged_usd': 10.0,
            'ebay_fee_usd': 10.0,
            'ebay_ad_fee_usd': 5.0,
        })

        # 生成 3 月份报表（无数据）
        output_path = tmp_path / "tax_report_2024-03.xlsx"
        result_path = generate_report(2024, str(output_path), month=3)

        assert os.path.exists(result_path)

        wb = load_workbook(output_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # 只有表头
        assert len(rows) == 1

        # 验证警告信息已打印
        captured = capsys.readouterr()
        assert "无订单数据" in captured.out or "⚠" in captured.out

    def test_generate_month_filename_format(self, test_db, tmp_path):
        """月份为 1 位数时补零"""
        insert('ebay_orders', {
            'order_id': 'ORDER-JAN',
            'sale_date': '2024-01-15',
            'buyer_username': 'buyer_jan',
            'item_title': 'January Item',
            'quantity': 1,
            'sale_price_usd': 100.0,
            'shipping_charged_usd': 10.0,
            'ebay_fee_usd': 10.0,
            'ebay_ad_fee_usd': 5.0,
        })

        # 生成 1 月份报表（month=1 不是 '01'）
        output_path = tmp_path / "tax_report_2024-01.xlsx"
        generate_report(2024, str(output_path), month=1)

        assert os.path.exists(output_path)

        wb = load_workbook(output_path)
        ws1 = wb["订单明细"]
        rows = list(ws1.iter_rows(values_only=True))

        # 验证包含 1 月订单
        assert len(rows) == 2
        assert rows[1][0] == 'ORDER-JAN'
