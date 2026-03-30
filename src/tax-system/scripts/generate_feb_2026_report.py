#!/usr/bin/env python3
"""
生成 2026 年 2 月 eBay 报税报表
用法：python3 generate_february_tax_report.py
"""
import sqlite3
import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 文件路径
EBAY_CSV = '/Users/arielhe/.openclaw/media/inbound/3e653725-419e-4e22-a37b-3cc75cb97bf3.csv'
AMAZON_CSV = '/Users/arielhe/.openclaw/media/inbound/de0d77eb-a9ce-404a-a815-bf97f1cbab5b.csv'
DB_PATH = '/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/orders.db'
OUTPUT_PATH = f'/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/outputs/2026-02/tax_report_2026-02_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

# 汇率
EXCHANGE_RATE = 150.0  # JPY/USD

def parse_currency(value):
    """解析货币字符串"""
    if not value:
        return 0.0
    value = str(value).replace('$', '').replace('¥', '').replace(',', '').replace('"', '').strip()
    try:
        return float(value)
    except:
        return 0.0

def parse_date_jp(value):
    """解析日本日期格式 (2026/03/22)"""
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), '%Y/%m/%d').strftime('%Y-%m-%d')
    except:
        return None

def parse_date_us(value):
    """解析美国日期格式 (Mar-15-26)"""
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), '%b-%d-%y').strftime('%Y-%m-%d')
    except:
        return None

def init_db():
    """初始化数据库表"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 创建表（如果不存在）
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ebay_orders (
            order_id TEXT PRIMARY KEY,
            sale_date DATE,
            buyer_username TEXT,
            item_title TEXT,
            item_id TEXT,
            quantity INTEGER DEFAULT 1,
            sale_price_usd REAL,
            shipping_charged_usd REAL,
            ebay_fee_usd REAL,
            ebay_ad_fee_usd REAL,
            payment_net_usd REAL,
            order_status TEXT,
            shipping_address_country TEXT,
            tracking_number TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS purchases (
            id TEXT PRIMARY KEY,
            platform TEXT NOT NULL,
            purchase_date DATE,
            item_name TEXT,
            item_name_en TEXT,
            item_sku TEXT,
            quantity INTEGER DEFAULT 1,
            unit_price_jpy REAL,
            total_price_jpy REAL,
            tax_jpy REAL,
            shipping_fee_jpy REAL,
            order_number TEXT,
            receipt_image_path TEXT,
            needs_review INTEGER DEFAULT 0,
            no_match_reason TEXT DEFAULT NULL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shipments (
            id TEXT PRIMARY KEY,
            carrier TEXT,
            tracking_number TEXT,
            ebay_order_id TEXT,
            ship_date DATE,
            shipping_fee_usd REAL,
            cpass_transaction_id TEXT,
            jp_post_email_path TEXT,
            match_method TEXT DEFAULT NULL,
            confirmed_by TEXT DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS purchase_order_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_id TEXT NOT NULL,
            ebay_order_id TEXT NOT NULL,
            match_method TEXT,
            confidence REAL,
            confirmed_by TEXT DEFAULT NULL,
            allocated_qty INTEGER DEFAULT NULL,
            allocated_cost_jpy REAL DEFAULT NULL,
            allocated_tax_jpy REAL DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inventory (
            id TEXT PRIMARY KEY,
            item_sku TEXT,
            item_name TEXT,
            item_name_en TEXT,
            total_quantity INTEGER,
            sold_quantity INTEGER DEFAULT 0,
            remaining_quantity INTEGER,
            total_cost_jpy REAL,
            total_tax_jpy REAL,
            average_cost_per_unit REAL,
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    conn.commit()
    conn.close()

def import_ebay_orders():
    """导入 eBay 订单"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    imported = 0
    feb_count = 0
    
    with open(EBAY_CSV, 'r', encoding='utf-8') as f:
        # 跳过空行，找到真正的表头
        lines = [line for line in f.readlines() if line.strip() and 'Order Number' in line]
        if not lines:
            print("  ⚠️ 未找到有效的 eBay 订单数据")
            return 0
        
        # 重新打开文件，跳过空行
        f.seek(0)
        non_empty_lines = [line for line in f if line.strip()]
        
        import io
        reader = csv.DictReader(io.StringIO(''.join(non_empty_lines)))
        
        for row in reader:
            order_number = row.get('Order Number', '').strip()
            if not order_number:
                continue
            
            sale_date = parse_date_us(row.get('Sale Date', ''))
            
            # 只导入 2 月份订单
            if sale_date and sale_date.startswith('2026-02'):
                feb_count += 1
            
            # 提取追踪号（从 Tracking Number 字段）
            tracking = row.get('Tracking Number', '').strip()
            
            # 解析金额
            sold_for = parse_currency(row.get('Sold For', '0'))
            shipping = parse_currency(row.get('Shipping And Handling', '0'))
            tax = parse_currency(row.get('eBay Collected Tax', '0'))
            total = parse_currency(row.get('Total Price', '0'))
            
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO ebay_orders (
                        order_id, sale_date, buyer_username, item_title, item_id,
                        quantity, sale_price_usd, shipping_charged_usd,
                        tracking_number, shipping_address_country
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_number,
                    sale_date,
                    row.get('Buyer Username', ''),
                    row.get('Item Title', ''),
                    row.get('Item Number', ''),
                    int(row.get('Quantity', 1) or 1),
                    sold_for,
                    shipping,
                    tracking,
                    row.get('Ship To Country', '')
                ))
                imported += 1
            except Exception as e:
                print(f"  跳过订单 {order_number}: {e}")
    
    conn.commit()
    conn.close()
    print(f"✅ 导入 eBay 订单：{imported} 条（2 月份：{feb_count} 条）")
    return feb_count

def import_amazon_purchases():
    """导入 Amazon 采购订单"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    imported = 0
    feb_count = 0
    
    with open(AMAZON_CSV, 'r', encoding='utf-8') as f:
        # 读取所有非空行
        lines = [line for line in f.readlines() if line.strip() and '注文日' in line]
        if not lines:
            print("  ⚠️ 未找到有效的 Amazon 采购数据")
            return 0
        
        # 重新打开文件
        f.seek(0)
        non_empty_lines = [line for line in f if line.strip() and '注文日' in line]
        
        import io
        reader = csv.DictReader(io.StringIO(''.join(non_empty_lines)))
        
        for row in reader:
            order_number = row.get('注文番号', '').strip()
            if not order_number:
                continue
            
            purchase_date = parse_date_jp(row.get('注文日', ''))
            
            # 只导入 2 月份采购
            if purchase_date and purchase_date.startswith('2026-02'):
                feb_count += 1
            
            # 生成唯一 ID
            asin = row.get('ASIN', '').strip()
            purchase_id = f"amazon_jp_{order_number}_{asin}" if asin else f"amazon_jp_{order_number}"
            
            # 解析金额
            total_with_tax = parse_currency(row.get('注文の合計（税込）', '0'))
            subtotal = parse_currency(row.get('注文の小計（税抜）', '0'))
            tax = parse_currency(row.get('注文の消費税額', '0'))
            shipping = parse_currency(row.get('注文の配送料および手数料（税抜）', '0'))
            quantity = int(row.get('注文の数量', 1) or 1)
            
            # 计算单价
            unit_price = subtotal / quantity if quantity > 0 else 0
            
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO purchases (
                        id, platform, purchase_date, item_name, item_sku,
                        quantity, unit_price_jpy, total_price_jpy, tax_jpy,
                        shipping_fee_jpy, order_number
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    purchase_id,
                    'amazon_jp',
                    purchase_date,
                    row.get('商品名', ''),
                    asin,
                    quantity,
                    unit_price,
                    subtotal,
                    tax,
                    shipping,
                    order_number
                ))
                imported += 1
            except Exception as e:
                print(f"  跳过采购 {order_number}: {e}")
    
    conn.commit()
    conn.close()
    print(f"✅ 导入 Amazon 采购：{imported} 条（2 月份：{feb_count} 条）")
    return feb_count

def run_matching():
    """运行匹配逻辑"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 简单的 ASIN 匹配
    cursor.execute("""
        INSERT OR IGNORE INTO purchase_order_links (
            purchase_id, ebay_order_id, match_method, confidence
        )
        SELECT p.id, e.order_id, 'asin', 1.0
        FROM purchases p
        JOIN ebay_orders e ON p.item_sku = e.item_id
        WHERE p.purchase_date LIKE '2026-02%'
          AND e.sale_date LIKE '2026-02%'
    """)
    
    matched = cursor.rowcount
    conn.commit()
    conn.close()
    print(f"✅ 匹配成功：{matched} 条")
    return matched

def generate_report(ebay_count, amazon_count, matched_count):
    """生成 Excel 报表"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    wb = Workbook()
    
    # 样式
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # === Sheet 1: 财务摘要 ===
    ws_summary = wb.active
    ws_summary.title = "财务摘要"
    
    ws_summary['A1'] = "2026 年 2 月 eBay 报税统计"
    ws_summary['A1'].font = Font(bold=True, size=16)
    
    # 查询统计数据
    cursor.execute("""
        SELECT 
            COUNT(*) as order_count,
            SUM(sale_price_usd) as total_sales,
            SUM(shipping_charged_usd) as total_shipping
        FROM ebay_orders 
        WHERE strftime('%Y-%m', sale_date) = '2026-02'
    """)
    stats = cursor.fetchone()
    
    order_count = stats['order_count'] or 0
    total_sales = stats['total_sales'] or 0
    total_shipping = stats['total_shipping'] or 0
    
    # 采购成本
    cursor.execute("""
        SELECT SUM(allocated_cost_jpy) as total FROM purchase_order_links
        WHERE ebay_order_id IN (
            SELECT order_id FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'
        )
    """)
    purchase_cost_jpy = cursor.fetchone()['total'] or 0
    purchase_cost_usd = purchase_cost_jpy / EXCHANGE_RATE
    
    # 净利润估算
    net_profit = total_sales + total_shipping - purchase_cost_usd
    
    summary_data = [
        ["项目", "数值"],
        ["订单数量", order_count],
        ["销售总额 (USD)", f"${total_sales:.2f}"],
        ["买家运费收入 (USD)", f"${total_shipping:.2f}"],
        ["采购成本 (JPY)", f"¥{purchase_cost_jpy:,.0f}"],
        ["采购成本 (USD)", f"${purchase_cost_usd:.2f}"],
        ["汇率", f"{EXCHANGE_RATE} JPY/USD"],
        ["", ""],
        ["净利润估算 (USD)", f"${net_profit:.2f}"],
        ["", ""],
        ["数据来源", ""],
        ["- eBay 订单", f"{ebay_count} 条"],
        ["- Amazon 采购", f"{amazon_count} 条"],
        ["- 匹配成功", f"{matched_count} 条"],
    ]
    
    for i, row in enumerate(summary_data, start=3):
        ws_summary.append(row)
    
    # === Sheet 2: 订单明细 ===
    ws_orders = wb.create_sheet(title="订单明细")
    
    order_headers = [
        "eBay 订单号", "成交日期", "商品名称", "数量", "销售价 (USD)",
        "买家运费 (USD)", "追踪号", "国家", "采购 ASIN", "采购成本 (JPY)",
        "采购成本 (USD)", "净利润 (USD)", "匹配状态"
    ]
    ws_orders.append(order_headers)
    
    for cell in ws_orders[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 查询 2 月份订单
    cursor.execute("""
        SELECT e.*, pol.purchase_id, p.total_price_jpy, p.item_sku
        FROM ebay_orders e
        LEFT JOIN purchase_order_links pol ON e.order_id = pol.ebay_order_id
        LEFT JOIN purchases p ON pol.purchase_id = p.id
        WHERE strftime('%Y-%m', e.sale_date) = '2026-02'
        ORDER BY e.sale_date
    """)
    
    for order in cursor.fetchall():
        purchase_cost = order['total_price_jpy'] or 0
        purchase_cost_usd = purchase_cost / EXCHANGE_RATE
        
        sale_price = order['sale_price_usd'] or 0
        shipping = order['shipping_charged_usd'] or 0
        
        net = sale_price + shipping - purchase_cost_usd if purchase_cost > 0 else 0
        
        match_status = "已匹配" if order['purchase_id'] else "未匹配"
        
        ws_orders.append([
            order['order_id'],
            order['sale_date'],
            (order['item_title'] or '')[:50],
            order['quantity'],
            sale_price,
            shipping,
            order['tracking_number'],
            order['shipping_address_country'],
            order['item_sku'],
            purchase_cost,
            round(purchase_cost_usd, 2),
            round(net, 2),
            match_status
        ])
    
    # === Sheet 3: 采购记录 ===
    ws_purchases = wb.create_sheet(title="采购记录")
    
    purchase_headers = [
        "采购 ID", "订单号", "采购日期", "商品名", "ASIN",
        "数量", "单价 (JPY)", "总价 (JPY)", "税额 (JPY)", "匹配状态"
    ]
    ws_purchases.append(purchase_headers)
    
    for cell in ws_purchases[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    cursor.execute("""
        SELECT p.*, pol.ebay_order_id
        FROM purchases p
        LEFT JOIN purchase_order_links pol ON p.id = pol.purchase_id
        WHERE strftime('%Y-%m', p.purchase_date) = '2026-02'
        ORDER BY p.purchase_date
    """)
    
    for p in cursor.fetchall():
        ws_purchases.append([
            p['id'],
            p['order_number'],
            p['purchase_date'],
            (p['item_name'] or '')[:50],
            p['item_sku'],
            p['quantity'],
            p['unit_price_jpy'],
            p['total_price_jpy'],
            p['tax_jpy'],
            '已匹配' if p['ebay_order_id'] else '未匹配'
        ])
    
    conn.close()
    
    # 保存文件
    import os
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    
    print(f"\n✅ 报表已生成：{OUTPUT_PATH}")
    print(f"\n📊 2 月份统计摘要:")
    print(f"   订单数量：{order_count}")
    print(f"   销售总额：${total_sales:.2f}")
    print(f"   采购成本：¥{purchase_cost_jpy:,.0f} (约 ${purchase_cost_usd:.2f})")
    print(f"   净利润估算：${net_profit:.2f}")

def main():
    print("🚀 开始生成 2026 年 2 月报税报表...\n")
    
    print("1️⃣ 初始化数据库...")
    init_db()
    
    print("\n2️⃣ 导入 eBay 订单...")
    ebay_count = import_ebay_orders()
    
    print("\n3️⃣ 导入 Amazon 采购...")
    amazon_count = import_amazon_purchases()
    
    print("\n4️⃣ 运行匹配...")
    matched_count = run_matching()
    
    print("\n5️⃣ 生成报表...")
    generate_report(ebay_count, amazon_count, matched_count)
    
    print("\n✅ 完成！")

if __name__ == '__main__':
    main()
