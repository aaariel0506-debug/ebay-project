#!/usr/bin/env python3
"""
生成 2026 年 2 月最终报税报表（含 Japan Post 运费）
"""
import sqlite3
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH = '/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/orders.db'
OUTPUT_DIR = '/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/outputs/2026-02'
OUTPUT_PATH = os.path.join(OUTPUT_DIR, 'tax_report_2026-02_FINAL.xlsx')
EXCHANGE_RATE = 150.0

BLUE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16)
CENTER = Alignment(horizontal="center", vertical="center")

def apply_header(ws, row_num=1):
    for cell in ws[row_num]:
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER

def clean_old_reports():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    old_files = glob.glob(os.path.join(OUTPUT_DIR, 'tax_report_2026-02_*.xlsx'))
    for f in old_files:
        if 'FINAL' not in f:
            os.remove(f)
            print(f"  🗑️ 删除：{os.path.basename(f)}")

def generate():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    wb = Workbook()
    
    # ========== Sheet 1: 财务摘要 ==========
    ws1 = wb.active
    ws1.title = "财务摘要"
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25
    
    ws1['A1'] = "📊 2026 年 2 月 eBay 报税统计"
    ws1['A1'].font = TITLE_FONT
    ws1['A2'] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    c.execute("SELECT COUNT(*) as cnt, SUM(sale_price_usd) as sales, SUM(shipping_charged_usd) as ship FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'")
    s = c.fetchone()
    orders = s['cnt'] or 0
    sales = s['sales'] or 0
    shipping = s['ship'] or 0
    
    c.execute("SELECT COUNT(*) as cnt FROM purchase_order_links")
    matched = c.fetchone()['cnt'] or 0
    
    c.execute("SELECT SUM(allocated_cost_jpy) as cost, SUM(allocated_tax_jpy) as tax FROM purchase_order_links")
    r = c.fetchone()
    matched_cost_jpy = r['cost'] or 0
    matched_tax_jpy = r['tax'] or 0
    matched_cost_usd = matched_cost_jpy / EXCHANGE_RATE
    
    c.execute("SELECT SUM(total_price_jpy) as cost FROM purchases WHERE strftime('%Y-%m', purchase_date) = '2026-02'")
    total_purchase_jpy = c.fetchone()['cost'] or 0
    
    c.execute("SELECT SUM(shipping_fee_jpy) as fee FROM japan_post_shipments WHERE matched = 1")
    jp_shipping_jpy = c.fetchone()['fee'] or 0
    jp_shipping_usd = jp_shipping_jpy / EXCHANGE_RATE
    
    total_cost_usd = matched_cost_usd + jp_shipping_usd
    profit = sales + shipping - total_cost_usd
    profit_margin = profit / (sales + shipping) * 100 if (sales + shipping) > 0 else 0
    match_rate = matched / orders * 100 if orders > 0 else 0
    
    data = [
        [],
        ["📦 销售数据", ""],
        ["eBay 订单数量", orders],
        ["销售总额 (USD)", sales],
        ["运费收入 (USD)", shipping],
        ["销售 + 运费合计 (USD)", sales + shipping],
        [],
        ["💰 成本数据", ""],
        ["采购匹配率", f"{matched}/{orders} ({match_rate:.0f}%)"],
        ["已匹配采购成本 (JPY)", matched_cost_jpy],
        ["已匹配采购税额 (JPY)", matched_tax_jpy],
        ["已匹配采购成本 (USD)", matched_cost_usd],
        ["Japan Post 运费 (JPY)", jp_shipping_jpy],
        ["Japan Post 运费 (USD)", jp_shipping_usd],
        ["总成本 (JPY)", matched_cost_jpy + jp_shipping_jpy],
        ["总成本 (USD)", total_cost_usd],
        ["2 月全部采购成本 (JPY)", total_purchase_jpy],
        [],
        ["📈 利润核算", ""],
        ["毛利润 (USD)", profit],
        ["利润率", f"{profit_margin:.1f}%"],
        ["汇率 (JPY/USD)", EXCHANGE_RATE],
        [],
        ["⚠️ 注意", ""],
        ["未包含", "eBay 平台费 (约 10-15%)"],
        ["未包含", "广告费"],
        ["未包含", "未匹配订单的采购成本"],
        ["已包含", "✅ Japan Post 运费"],
    ]
    
    for row in data:
        ws1.append(row)
    
    # ========== Sheet 2: 订单明细 ==========
    ws2 = wb.create_sheet("订单明细")
    headers2 = ["eBay 订单号", "日期", "买家", "商品名称", "数量", "售价 USD", "运费 USD", 
                "追踪号", "国家", "Amazon 订单号", "采购商品", "采购成本 JPY", "采购成本 USD", 
                "毛利润 USD", "匹配状态", "匹配分数"]
    ws2.append(headers2)
    apply_header(ws2)
    
    c.execute("""
        SELECT e.*, pol.purchase_id, pol.confidence, pol.allocated_cost_jpy,
               p.item_name as purchase_name, p.order_number as amazon_order
        FROM ebay_orders e
        LEFT JOIN purchase_order_links pol ON e.order_id = pol.ebay_order_id
        LEFT JOIN purchases p ON pol.purchase_id = p.id
        WHERE strftime('%Y-%m', e.sale_date) = '2026-02'
        ORDER BY e.sale_date
    """)
    
    row_num = 2
    for o in c.fetchall():
        cost_jpy = o['allocated_cost_jpy'] or 0
        cost_usd = cost_jpy / EXCHANGE_RATE
        sale = (o['sale_price_usd'] or 0) + (o['shipping_charged_usd'] or 0)
        profit_val = sale - cost_usd if cost_jpy > 0 else ''
        status = "✅ 已匹配" if o['purchase_id'] else "❌ 未匹配"
        score = int((o['confidence'] or 0) * 100) if o['confidence'] else ''
        
        ws2.append([
            o['order_id'], o['sale_date'], o['buyer_username'],
            (o['item_title'] or '')[:50], o['quantity'],
            o['sale_price_usd'], o['shipping_charged_usd'],
            o['tracking_number'], o['shipping_address_country'],
            o['amazon_order'] or '',
            (o['purchase_name'] or '')[:40] if o['purchase_name'] else '',
            cost_jpy if cost_jpy > 0 else '',
            round(cost_usd, 2) if cost_jpy > 0 else '',
            round(profit_val, 2) if isinstance(profit_val, float) else '',
            status, score
        ])
        
        status_cell = ws2.cell(row=row_num, column=14)
        if o['purchase_id']:
            status_cell.fill = PatternFill(start_color="D5F5D5", end_color="D5F5D5", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        row_num += 1
    
    # ========== Sheet 3: 采购记录 ==========
    ws3 = wb.create_sheet("采购记录（2 月）")
    headers3 = ["采购 ID", "订单号", "日期", "商品名称", "ASIN", "数量", "总价 JPY", "税额 JPY", "匹配 eBay 订单"]
    ws3.append(headers3)
    apply_header(ws3)
    
    c.execute("""
        SELECT p.*, pol.ebay_order_id
        FROM purchases p
        LEFT JOIN purchase_order_links pol ON p.id = pol.purchase_id
        WHERE strftime('%Y-%m', p.purchase_date) = '2026-02'
        ORDER BY p.purchase_date
    """)
    
    for p in c.fetchall():
        ws3.append([
            p['id'], p['order_number'], p['purchase_date'],
            (p['item_name'] or '')[:50], p['item_sku'],
            p['quantity'], p['total_price_jpy'], p['tax_jpy'],
            p['ebay_order_id'] or ''
        ])
    
    # ========== Sheet 4: Japan Post 运费 ==========
    ws4 = wb.create_sheet("Japan Post 运费（2 月）")
    headers4 = ["快递单号", "日期", "收件人", "目的地", "运费 JPY", "重量 g", "eBay 订单号", "匹配状态"]
    ws4.append(headers4)
    apply_header(ws4)
    
    c.execute("""
        SELECT tracking_number, ship_date, recipient, country, shipping_fee_jpy, weight_g, ebay_order_id, matched
        FROM japan_post_shipments
        ORDER BY ship_date
    """)
    
    for s in c.fetchall():
        status = "✅ 已匹配" if s['ebay_order_id'] else "❌ 未匹配"
        ws4.append([
            s['tracking_number'], s['ship_date'], s['recipient'] or '',
            s['country'] or '', s['shipping_fee_jpy'], s['weight_g'],
            s['ebay_order_id'] or '', status
        ])
    
    # ========== Sheet 5: 匹配详情 ==========
    ws5 = wb.create_sheet("匹配详情")
    headers5 = ["eBay 订单号", "eBay 商品", "Amazon 订单号", "Amazon 商品", "匹配分数", "采购成本 JPY"]
    ws5.append(headers5)
    apply_header(ws5)
    
    c.execute("""
        SELECT pol.*, e.item_title as ebay_title, p.item_name as amazon_name, p.order_number as amazon_order
        FROM purchase_order_links pol
        JOIN ebay_orders e ON pol.ebay_order_id = e.order_id
        JOIN purchases p ON pol.purchase_id = p.id
        ORDER BY pol.confidence DESC
    """)
    
    for m in c.fetchall():
        ws5.append([
            m['ebay_order_id'],
            (m['ebay_title'] or '')[:50],
            m['amazon_order'] or '',
            (m['amazon_name'] or '')[:50],
            int((m['confidence'] or 0) * 100),
            m['allocated_cost_jpy']
        ])
    
    wb.save(OUTPUT_PATH)
    conn.close()
    
    print(f"\n✅ 最终报表：{OUTPUT_PATH}")
    print(f"\n{'='*60}")
    print(f"📊 2026 年 2 月完整成本核算")
    print(f"{'='*60}")
    print(f"  订单数量：{orders}")
    print(f"  销售总额：${sales:.2f}")
    print(f"  运费收入：${shipping:.2f}")
    print(f"  销售 + 运费合计：${sales + shipping:.2f}")
    print(f"")
    print(f"  采购成本：¥{matched_cost_jpy:,.0f} (${matched_cost_usd:.2f})")
    print(f"  Japan Post 运费：¥{jp_shipping_jpy:,.0f} (${jp_shipping_usd:.2f})")
    print(f"  总成本：¥{matched_cost_jpy + jp_shipping_jpy:,.0f} (${total_cost_usd:.2f})")
    print(f"")
    print(f"  💰 毛利润：${profit:.2f}")
    print(f"  利润率：{profit_margin:.1f}%")
    print(f"{'='*60}")

print("🚀 生成最终报表（含 Japan Post 运费）...\n")
print("1️⃣ 清理旧报表...")
clean_old_reports()
print("\n2️⃣ 生成最终版...")
generate()
print("\n✅ 完成！")
