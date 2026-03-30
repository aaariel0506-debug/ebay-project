#!/usr/bin/env python3
"""
生成 2026 年 2 月 eBay 报税报表 - 修复版
"""
import sqlite3
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

EBAY_CSV = '/Users/arielhe/.openclaw/media/inbound/3e653725-419e-4e22-a37b-3cc75cb97bf3.csv'
AMAZON_CSV = '/Users/arielhe/.openclaw/media/inbound/de0d77eb-a9ce-404a-a815-bf97f1cbab5b.csv'
DB_PATH = '/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/orders.db'
OUTPUT_PATH = f'/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/outputs/2026-02/tax_report_2026-02_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
EXCHANGE_RATE = 150.0

def parse_money(val):
    try:
        return float(str(val).replace('$','').replace('¥','').replace(',','').replace('"','').strip())
    except:
        return 0.0

def parse_date_us(val):
    try:
        return datetime.strptime(str(val).strip(), '%b-%d-%y').strftime('%Y-%m-%d')
    except:
        return None

def parse_date_jp(val):
    try:
        return datetime.strptime(str(val).strip(), '%Y/%m/%d').strftime('%Y-%m-%d')
    except:
        return None

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS ebay_orders (order_id TEXT PRIMARY KEY, sale_date DATE, buyer_username TEXT, item_title TEXT, item_id TEXT, quantity INTEGER, sale_price_usd REAL, shipping_charged_usd REAL, tracking_number TEXT, shipping_address_country TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS purchases (id TEXT PRIMARY KEY, platform TEXT, purchase_date DATE, item_name TEXT, item_sku TEXT, quantity INTEGER, total_price_jpy REAL, tax_jpy REAL, order_number TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS purchase_order_links (purchase_id TEXT, ebay_order_id TEXT, match_method TEXT, confidence REAL)")
    conn.commit()
    conn.close()

def import_ebay():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    imported = 0
    feb = 0
    
    with open(EBAY_CSV, 'r', encoding='utf-8') as f:
        lines = [l for l in f.readlines() if l.strip()]
    
    # 从行 3 开始（跳过头 2 行）
    for line in lines[3:]:
        try:
            row = list(csv.reader([line]))[0]
            if len(row) < 50:
                continue
            
            order_num = row[1].strip()  # Order Number
            if not order_num:
                continue
            
            sale_date = parse_date_us(row[49]) if len(row) > 49 else None  # Sale Date
            if sale_date and sale_date.startswith('2026-02'):
                feb += 1
            
            c.execute("INSERT OR REPLACE INTO ebay_orders VALUES (?,?,?,?,?,?,?,?,?,?)",
                (order_num, sale_date, row[2], row[22], row[21], 
                 int(row[25]) if row[25].strip() else 1,
                 parse_money(row[26]), parse_money(row[27]),
                 row[62] if len(row) > 62 else '', row[20]))
            imported += 1
        except Exception as e:
            pass
    
    conn.commit()
    conn.close()
    print(f"✅ eBay 订单：{imported} 条（2 月：{feb} 条）")
    return feb

def import_amazon():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    imported = 0
    feb = 0
    
    with open(AMAZON_CSV, 'r', encoding='utf-8') as f:
        lines = [l for l in f.readlines() if l.strip() and '注文日' not in l]
    
    for line in lines:
        try:
            row = list(csv.reader([line]))[0]
            if len(row) < 10:
                continue
            
            order_num = row[1].strip() if len(row) > 1 else ''
            if not order_num:
                continue
            
            purchase_date = parse_date_jp(row[0]) if len(row) > 0 else None
            if purchase_date and purchase_date.startswith('2026-02'):
                feb += 1
            
            asin = row[52].strip() if len(row) > 52 else ''
            pid = f"amazon_{order_num}_{asin}" if asin else f"amazon_{order_num}"
            
            c.execute("INSERT OR REPLACE INTO purchases VALUES (?,?,?,?,?,?,?,?,?)",
                (pid, 'amazon_jp', purchase_date, row[53] if len(row) > 53 else '',
                 asin, int(row[3]) if row[3].strip() else 1,
                 parse_money(row[6]), parse_money(row[8]), order_num))
            imported += 1
        except Exception as e:
            pass
    
    conn.commit()
    conn.close()
    print(f"✅ Amazon 采购：{imported} 条（2 月：{feb} 条）")
    return feb

def match():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM purchase_order_links")
    c.execute("""INSERT INTO purchase_order_links SELECT p.id, e.order_id, 'asin', 1.0 
                 FROM purchases p JOIN ebay_orders e ON p.item_sku = e.item_id
                 WHERE p.purchase_date LIKE '2026-02%' AND e.sale_date LIKE '2026-02%'""")
    matched = c.rowcount
    conn.commit()
    conn.close()
    print(f"✅ 匹配：{matched} 条")
    return matched

def report(ebay_count, amazon_count, matched_count):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    wb = Workbook()
    ws = wb.active
    ws.title = "财务摘要"
    
    c.execute("SELECT COUNT(*) as cnt, SUM(sale_price_usd) as sales, SUM(shipping_charged_usd) as ship FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'")
    s = c.fetchone()
    
    orders = s['cnt'] or 0
    sales = s['sales'] or 0
    shipping = s['ship'] or 0
    
    c.execute("SELECT SUM(total_price_jpy) as cost FROM purchases WHERE strftime('%Y-%m', purchase_date) = '2026-02'")
    cost_jpy = c.fetchone()['cost'] or 0
    cost_usd = cost_jpy / EXCHANGE_RATE
    
    profit = sales + shipping - cost_usd
    
    ws['A1'] = "2026 年 2 月 eBay 报税统计"
    ws['A1'].font = Font(bold=True, size=16)
    
    data = [
        ["项目", "数值"],
        ["订单数量", orders],
        ["销售总额 (USD)", f"${sales:.2f}"],
        ["运费收入 (USD)", f"${shipping:.2f}"],
        ["采购成本 (JPY)", f"¥{cost_jpy:,.0f}"],
        ["采购成本 (USD)", f"${cost_usd:.2f}"],
        ["汇率", f"{EXCHANGE_RATE} JPY/USD"],
        ["", ""],
        ["净利润估算 (USD)", f"${profit:.2f}"],
        ["", ""],
        ["数据来源", ""],
        ["- eBay 订单", f"{ebay_count} 条"],
        ["- Amazon 采购", f"{amazon_count} 条"],
        ["- 匹配成功", f"{matched_count} 条"],
    ]
    
    for i, row in enumerate(data, start=3):
        ws.append(row)
    
    # 订单明细
    ws2 = wb.create_sheet("订单明细")
    ws2.append(["订单号", "日期", "商品", "数量", "售价", "运费", "追踪号", "国家", "利润"])
    c.execute("SELECT * FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02' ORDER BY sale_date")
    for o in c.fetchall():
        ws2.append([o['order_id'], o['sale_date'], (o['item_title'] or '')[:40], o['quantity'], o['sale_price_usd'], o['shipping_charged_usd'], o['tracking_number'], o['shipping_address_country'], ''])
    
    # 采购记录
    ws3 = wb.create_sheet("采购记录")
    ws3.append(["采购 ID", "订单号", "日期", "商品", "ASIN", "数量", "总价 JPY", "税额"])
    c.execute("SELECT * FROM purchases WHERE strftime('%Y-%m', purchase_date) = '2026-02' ORDER BY purchase_date")
    for p in c.fetchall():
        ws3.append([p['id'], p['order_number'], p['purchase_date'], (p['item_name'] or '')[:40], p['item_sku'], p['quantity'], p['total_price_jpy'], p['tax_jpy']])
    
    import os
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    conn.close()
    
    print(f"\n✅ 报表：{OUTPUT_PATH}")
    print(f"\n📊 2 月统计:")
    print(f"   订单：{orders}")
    print(f"   销售：${sales:.2f}")
    print(f"   采购：¥{cost_jpy:,.0f}")
    print(f"   利润：${profit:.2f}")

print("🚀 开始生成 2 月报表...\n")
init_db()
eb = import_ebay()
am = import_amazon()
mt = match()
report(eb, am, mt)
print("\n✅ 完成！")
