#!/usr/bin/env python3
"""
生成 2026 年 2 月 eBay 报税报表 - 最终版
"""
import sqlite3
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

EBAY_CSV = '/Users/arielhe/.openclaw/media/inbound/3e653725-419e-4e22-a37b-3cc75cb97bf3.csv'
AMAZON_CSV = '/Users/arielhe/.openclaw/media/inbound/de0d77eb-a9ce-404a-a815-bf97f1cbab5b.csv'
DB_PATH = '/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/orders.db'
OUTPUT_PATH = f'/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/outputs/2026-02/tax_report_2026-02_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
EXCHANGE_RATE = 150.0

def parse_money(val):
    try:
        return float(str(val).replace('$','').replace('¥','').replace(',','').replace('"','').strip())
    except:
        return 0.0

def parse_date_us(val):
    try:
        return datetime.strptime(str(val).strip(), '%b-%d-%y').strftime('%Y-%m-%d')
    except:
        return None

def parse_date_jp(val):
    try:
        return datetime.strptime(str(val).strip(), '%Y/%m/%d').strftime('%Y-%m-%d')
    except:
        return None

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS ebay_orders")
    c.execute("DROP TABLE IF EXISTS purchases")
    c.execute("DROP TABLE IF EXISTS purchase_order_links")
    
    c.execute("""CREATE TABLE ebay_orders (
        order_id TEXT PRIMARY KEY, sale_date DATE, buyer_username TEXT, item_title TEXT, 
        item_id TEXT, quantity INTEGER, sale_price_usd REAL, shipping_charged_usd REAL, 
        tracking_number TEXT, shipping_address_country TEXT)""")
    
    c.execute("""CREATE TABLE purchases (
        id TEXT PRIMARY KEY, platform TEXT, purchase_date DATE, item_name TEXT, 
        item_sku TEXT, quantity INTEGER, total_price_jpy REAL, tax_jpy REAL, order_number TEXT)""")
    
    c.execute("""CREATE TABLE purchase_order_links (
        id INTEGER PRIMARY KEY AUTOINCREMENT, purchase_id TEXT, ebay_order_id TEXT, 
        match_method TEXT, confidence REAL, allocated_qty INTEGER, 
        allocated_cost_jpy REAL, allocated_tax_jpy REAL)""")
    
    conn.commit()
    conn.close()

def import_ebay():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    imported = 0
    feb = 0
    
    with open(EBAY_CSV, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # 找表头行（包含"Order Number"的行）
    header_idx = 0
    for i, row in enumerate(rows):
        if len(row) > 1 and 'Order Number' in row[1]:
            header_idx = i
            break
    
    header = rows[header_idx]
    print(f"  表头行：{header_idx}, 列数：{len(header)}")
    
    # 找列索引
    col_map = {}
    for i, col in enumerate(header):
        cl = col.lower()
        if 'order number' in cl: col_map['order'] = i
        elif 'sale date' in cl: col_map['date'] = i
        elif 'sold for' in cl: col_map['sold_for'] = i
        elif 'shipping and handling' in cl: col_map['shipping'] = i
        elif 'tracking number' in cl: col_map['tracking'] = i
        elif 'ship to country' in cl: col_map['country'] = i
        elif 'item title' in cl: col_map['title'] = i
        elif 'item number' in cl: col_map['item_id'] = i
        elif 'quantity' in cl and 'promoted' not in cl: col_map['qty'] = i
        elif 'buyer username' in cl: col_map['buyer'] = i
    
    print(f"  列映射：{col_map}")
    
    # 跳过空行和表头
    data_start = header_idx + 1
    while data_start < len(rows) and (not rows[data_start] or not rows[data_start][0].strip()):
        data_start += 1
    
    for row in rows[data_start:]:
        if len(row) < 5:
            continue
        
        order_num = row[col_map.get('order', 1)].strip() if col_map.get('order', 1) < len(row) else ''
        if not order_num:
            continue
        
        date_idx = col_map.get('date', -1)
        sale_date = parse_date_us(row[date_idx]) if date_idx >= 0 and date_idx < len(row) else None
        
        if sale_date and sale_date.startswith('2026-02'):
            feb += 1
        
        sold_idx = col_map.get('sold_for', 26)
        ship_idx = col_map.get('shipping', 27)
        sold_for = parse_money(row[sold_idx]) if sold_idx < len(row) else 0
        shipping = parse_money(row[ship_idx]) if ship_idx < len(row) else 0
        
        try:
            c.execute("INSERT INTO ebay_orders VALUES (?,?,?,?,?,?,?,?,?,?)",
                (order_num, sale_date,
                 row[col_map.get('buyer', 2)] if col_map.get('buyer', 2) < len(row) else '',
                 row[col_map.get('title', 22)] if col_map.get('title', 22) < len(row) else '',
                 row[col_map.get('item_id', 21)] if col_map.get('item_id', 21) < len(row) else '',
                 int(row[col_map.get('qty', 25)]) if col_map.get('qty', 25) < len(row) and row[col_map.get('qty', 25)].strip() else 1,
                 sold_for, shipping,
                 row[col_map.get('tracking', 62)] if col_map.get('tracking', 62) < len(row) else '',
                 row[col_map.get('country', 20)] if col_map.get('country', 20) < len(row) else ''))
            imported += 1
        except Exception as e:
            print(f"  跳过 {order_num}: {e}")
    
    conn.commit()
    conn.close()
    print(f"✅ eBay 订单：{imported} 条（2 月：{feb} 条）")
    return feb

def import_amazon():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    imported = 0
    feb = 0
    
    with open(AMAZON_CSV, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        
        col_map = {}
        for i, col in enumerate(header):
            if '注文番号' in col: col_map['order'] = i
            elif '注文日' in col: col_map['date'] = i
            elif '商品名' in col: col_map['name'] = i
            elif 'ASIN' in col: col_map['asin'] = i
            elif '注文の数量' in col: col_map['qty'] = i
            elif '注文の合計（税込）' in col: col_map['total'] = i
            elif '注文の消費税額' in col: col_map['tax'] = i
        
        print(f"  Amazon 列映射：{col_map}")
        
        for row in reader:
            if len(row) < 5:
                continue
            
            order_num = row[col_map.get('order', 1)].strip() if col_map.get('order', 1) < len(row) else ''
            if not order_num:
                continue
            
            date_idx = col_map.get('date', 0)
            purchase_date = parse_date_jp(row[date_idx]) if date_idx < len(row) else None
            
            if purchase_date and purchase_date.startswith('2026-02'):
                feb += 1
            
            asin_idx = col_map.get('asin', 52)
            asin = row[asin_idx].strip() if asin_idx < len(row) else ''
            pid = f"amazon_{order_num}_{asin}" if asin else f"amazon_{order_num}"
            
            qty_idx = col_map.get('qty', 3)
            total_idx = col_map.get('total', 6)
            tax_idx = col_map.get('tax', 8)
            name_idx = col_map.get('name', 53)
            
            qty = int(row[qty_idx]) if qty_idx < len(row) and row[qty_idx].strip() else 1
            total = parse_money(row[total_idx]) if total_idx < len(row) else 0
            tax = parse_money(row[tax_idx]) if tax_idx < len(row) else 0
            name = row[name_idx] if name_idx < len(row) else ''
            
            try:
                c.execute("INSERT INTO purchases VALUES (?,?,?,?,?,?,?,?,?)",
                    (pid, 'amazon_jp', purchase_date, name, asin, qty, total, tax, order_num))
                imported += 1
            except Exception as e:
                pass
    
    conn.commit()
    conn.close()
    print(f"✅ Amazon 采购：{imported} 条（2 月：{feb} 条）")
    return feb

def match():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM purchase_order_links")
    c.execute("""INSERT INTO purchase_order_links (purchase_id, ebay_order_id, match_method, confidence)
                 SELECT p.id, e.order_id, 'asin', 1.0 
                 FROM purchases p 
                 JOIN ebay_orders e ON p.item_sku = e.item_id
                 WHERE p.purchase_date LIKE '2026-02%' AND e.sale_date LIKE '2026-02%'""")
    matched = c.rowcount
    conn.commit()
    conn.close()
    print(f"✅ 匹配：{matched} 条")
    return matched

def report(eb_count, am_count, mt_count):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    wb = Workbook()
    ws = wb.active
    ws.title = "财务摘要"
    
    c.execute("SELECT COUNT(*) as cnt, SUM(sale_price_usd) as sales, SUM(shipping_charged_usd) as ship FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'")
    s = c.fetchone()
    orders = s['cnt'] or 0
    sales = s['sales'] or 0
    shipping = s['ship'] or 0
    
    c.execute("SELECT SUM(total_price_jpy) as cost FROM purchases WHERE strftime('%Y-%m', purchase_date) = '2026-02'")
    cost_jpy = c.fetchone()['cost'] or 0
    cost_usd = cost_jpy / EXCHANGE_RATE
    profit = sales + shipping - cost_usd
    
    ws['A1'] = "2026 年 2 月 eBay 报税统计"
    ws['A1'].font = Font(bold=True, size=16)
    
    data = [
        ["项目", "数值"],
        ["订单数量", orders],
        ["销售总额 (USD)", f"${sales:.2f}"],
        ["运费收入 (USD)", f"${shipping:.2f}"],
        ["采购成本 (JPY)", f"¥{cost_jpy:,.0f}"],
        ["采购成本 (USD)", f"${cost_usd:.2f}"],
        ["汇率", f"{EXCHANGE_RATE} JPY/USD"],
        ["", ""],
        ["净利润估算 (USD)", f"${profit:.2f}"],
        ["", ""],
        ["数据来源", ""],
        ["- eBay 订单", f"{eb_count} 条"],
        ["- Amazon 采购", f"{am_count} 条"],
        ["- 匹配成功", f"{mt_count} 条"],
    ]
    for row in data:
        ws.append(row)
    
    ws2 = wb.create_sheet("订单明细")
    ws2.append(["订单号", "日期", "商品", "数量", "售价", "运费", "追踪号", "国家"])
    c.execute("SELECT * FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02' ORDER BY sale_date")
    for o in c.fetchall():
        ws2.append([o['order_id'], o['sale_date'], (o['item_title'] or '')[:40], o['quantity'], o['sale_price_usd'], o['shipping_charged_usd'], o['tracking_number'], o['shipping_address_country']])
    
    ws3 = wb.create_sheet("采购记录")
    ws3.append(["采购 ID", "订单号", "日期", "商品", "ASIN", "数量", "总价 JPY", "税额"])
    c.execute("SELECT * FROM purchases WHERE strftime('%Y-%m', purchase_date) = '2026-02' ORDER BY purchase_date")
    for p in c.fetchall():
        ws3.append([p['id'], p['order_number'], p['purchase_date'], (p['item_name'] or '')[:40], p['item_sku'], p['quantity'], p['total_price_jpy'], p['tax_jpy']])
    
    import os
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    conn.close()
    
    print(f"\n✅ 报表：{OUTPUT_PATH}")
    print(f"\n📊 2 月统计:")
    print(f"   订单：{orders}")
    print(f"   销售：${sales:.2f}")
    print(f"   采购：¥{cost_jpy:,.0f} (约 ${cost_usd:.2f})")
    print(f"   利润：${profit:.2f}")

print("🚀 开始生成 2 月报表...\n")
init_db()
eb = import_ebay()
am = import_amazon()
mt = match()
report(eb, am, mt)
print("\n✅ 完成！")
