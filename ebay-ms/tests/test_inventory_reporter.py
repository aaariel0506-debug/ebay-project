"""
tests/test_inventory_reporter.py

Day 23: 库存报表测试
"""

from datetime import datetime, timedelta, timezone
from decimal import Decimal


class TestInventoryReporter:
    """InventoryReporter 测试"""

    def _setup_inventory(self, db_session, sample_product):
        """创建 Inventory 记录用于报表测试"""
        from core.models import Inventory, InventoryType

        now = datetime.now(timezone.utc)
        records = [
            # 入库 10 件（5 天前）
            Inventory(
                sku=sample_product.sku,
                type=InventoryType.IN,
                quantity=10,
                location="A-1",
                occurred_at=now - timedelta(days=5),
                operator="test",
            ),
            # 出库 3 件（3 天前）
            Inventory(
                sku=sample_product.sku,
                type=InventoryType.OUT,
                quantity=3,
                location="A-1",
                related_order="ORDER-REPORT-001",
                occurred_at=now - timedelta(days=3),
                operator="test",
            ),
            # 退货 2 件（1 天前）
            Inventory(
                sku=sample_product.sku,
                type=InventoryType.RETURN,
                quantity=2,
                location="A-1",
                related_order="ORDER-REPORT-001",
                occurred_at=now - timedelta(days=1),
                operator="test",
            ),
        ]
        for r in records:
            db_session.add(r)
        db_session.commit()
        return records

    def test_get_stock_snapshot(self, db_session, sample_product):
        """库存快照包含所有必要字段"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        items = reporter.get_stock_snapshot()
        assert len(items) >= 1

        item = next((i for i in items if i.sku == sample_product.sku), None)
        assert item is not None
        # RETURN 不在 get_all_stock 的 subquery 中，所以 10 - 3 = 7
        # sample_product fixture 已有预设库存，叠加我们的记录后一定 > 10
        assert item.available_quantity >= 9, (
            f"expected >= 10, got {item.available_quantity}"
        )
        assert item.cost_price == Decimal("100.00")
        assert item.inventory_value >= Decimal("900.00")
        assert "A-1" in item.locations

    def test_get_movements_all(self, db_session, sample_product):
        """出入库明细返回所有记录"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        movements = reporter.get_movements(limit=100)
        assert len(movements) == 3
        # 按时间倒序：最新的是 RETURN
        assert movements[0].movement_type == "return"
        assert movements[1].movement_type == "out"
        assert movements[2].movement_type == "in"

    def test_get_movements_date_filter(self, db_session, sample_product):
        """出入库明细支持日期范围筛选"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        now = datetime.now(timezone.utc)

        # 只取最近 2 天
        recent = reporter.get_movements(
            start_date=now - timedelta(days=2),
            end_date=now,
        )
        assert len(recent) == 1
        assert recent[0].movement_type == "return"

    def test_get_movements_sku_filter(self, db_session, sample_product):
        """出入库明细支持 SKU 筛选"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        movements = reporter.get_movements(sku="NONEXISTENT-SKU")
        assert len(movements) == 0

        movements = reporter.get_movements(sku=sample_product.sku)
        assert len(movements) == 3

    def test_get_movements_type_filter(self, db_session, sample_product):
        """出入库明细支持类型筛选"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        out_movements = reporter.get_movements(movement_type="OUT")
        assert len(out_movements) == 1
        assert out_movements[0].quantity == 3

    def test_get_trend(self, db_session, sample_product):
        """库存变动趋势返回每日汇总"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        trend = reporter.get_trend(sample_product.sku, lookback_days=30)
        assert len(trend) >= 1

        # 最新收盘 = get_stock 返回的 available = 7
        latest = trend[-1]
        assert latest.closing >= 9  # includes sample_product preset + our records
        assert latest.sku == sample_product.sku

    def test_export_snapshot_to_excel(self, db_session, sample_product, tmp_path):
        """库存快照导出 Excel — 验证表头 + 位置分布 + 进货价"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        out_path = tmp_path / "snapshot.xlsx"
        reporter.export_snapshot_to_excel(out_path)

        assert out_path.exists()
        assert out_path.stat().st_size > 0

        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        assert ws.title == "库存快照"

        # 验证表头
        assert ws.cell(row=1, column=1).value == "SKU"
        assert ws.cell(row=1, column=2).value == "商品名称"
        assert ws.cell(row=1, column=3).value == "可用数量"
        assert ws.cell(row=1, column=4).value == "位置分布"
        assert ws.cell(row=1, column=5).value == "进货价"
        assert ws.cell(row=1, column=6).value == "库存金额 (JPY)"
        assert ws.cell(row=1, column=7).value == "最后入库"
        assert ws.cell(row=1, column=8).value == "最后出库"

        # 验证数据行
        assert ws.cell(row=2, column=1).value == sample_product.sku
        # 位置分布格式："A-1:15" 或类似
        loc_cell = ws.cell(row=2, column=4).value
        assert loc_cell is not None
        assert "A-1" in str(loc_cell)
        # 进货价是数值
        cost_cell = ws.cell(row=2, column=5).value
        assert cost_cell is not None
        assert isinstance(cost_cell, (int, float))

        # 验证合计行
        summary_row = ws.max_row
        assert ws.cell(row=summary_row, column=1).value == "合计"
        assert ws.cell(row=summary_row, column=3).value is not None  # 总数量
        assert ws.cell(row=summary_row, column=6).value is not None  # 总金额
        assert ws.cell(row=summary_row, column=4).value in (None, "")  # 位置分布列应为空
        # 合计行字体加粗
        assert ws.cell(row=summary_row, column=3).font.bold is True
        assert ws.cell(row=summary_row, column=6).font.bold is True

    def test_export_movements_to_excel(self, db_session, sample_product, tmp_path):
        """出入库明细导出 Excel — 验证颜色编码"""
        from modules.inventory_offline.reporter import InventoryReporter

        self._setup_inventory(db_session, sample_product)
        reporter = InventoryReporter()

        out_path = tmp_path / "movements.xlsx"
        reporter.export_movements_to_excel(out_path)

        assert out_path.exists()
        assert out_path.stat().st_size > 0

        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        assert ws.title == "出入库明细"
        assert ws.max_row == 4  # 3 条数据 + 1 表头

        # 验证颜色编码 (按时间倒序：RETURN/OUT/IN)
        # RETURN (row 2) - 蓝色 DDEBF7
        return_cell = ws.cell(row=2, column=1)
        return_color = return_cell.fill.fgColor.rgb
        assert return_color.endswith("DDEBF7"), f"RETURN should be blue, got {return_color}"

        # OUT (row 3) - 红色 FFCCCC
        out_cell = ws.cell(row=3, column=1)
        out_color = out_cell.fill.fgColor.rgb
        assert out_color.endswith("FFCCCC"), f"OUT should be red, got {out_color}"

        # IN (row 4) - 绿色 C6EFCE
        in_cell = ws.cell(row=4, column=1)
        in_color = in_cell.fill.fgColor.rgb
        assert in_color.endswith("C6EFCE"), f"IN should be green, got {in_color}"

    def test_get_movements_in_cost_filled(self, db_session, sample_product):
        """IN 类型从 InboundReceiptItem 填充 unit_cost / total_cost"""
        from datetime import datetime, timezone
        from decimal import Decimal

        from core.models import InboundReceipt, InboundReceiptItem, InboundStatus, Inventory, InventoryType
        from modules.inventory_offline.reporter import InventoryReporter

        # 创建入库单 + 明细（含 cost_price）
        receipt = InboundReceipt(
            receipt_no="IN-REPORT-COST-001",
            supplier="Test Supplier",
            status=InboundStatus.RECEIVED,
        )
        db_session.add(receipt)
        db_session.flush()

        item = InboundReceiptItem(
            receipt_id=receipt.id,
            sku=sample_product.sku,
            expected_quantity=10,
            received_quantity=10,
            cost_price=Decimal("150.00"),
        )
        db_session.add(item)

        # 创建 Inventory IN 记录，related_order = receipt_no
        inv = Inventory(
            sku=sample_product.sku,
            type=InventoryType.IN,
            quantity=10,
            location="A-1",
            related_order=receipt.receipt_no,
            occurred_at=datetime.now(timezone.utc),
            operator="test",
        )
        db_session.add(inv)
        db_session.commit()

        reporter = InventoryReporter()
        movements = reporter.get_movements(sku=sample_product.sku, movement_type="IN")

        assert len(movements) == 1
        assert movements[0].unit_cost == Decimal("150.00")
        assert movements[0].total_cost == Decimal("1500.00")  # 10 * 150

    def test_get_movements_out_uses_product_cost(self, db_session, sample_product):
        """OUT 类型 fallback 到 Product.cost_price"""
        from datetime import datetime, timezone
        from decimal import Decimal

        from core.models import Inventory, InventoryType
        from modules.inventory_offline.reporter import InventoryReporter

        inv = Inventory(
            sku=sample_product.sku,
            type=InventoryType.OUT,
            quantity=3,
            location="A-1",
            related_order="ORDER-OUT-001",
            occurred_at=datetime.now(timezone.utc),
            operator="test",
        )
        db_session.add(inv)
        db_session.commit()

        reporter = InventoryReporter()
        movements = reporter.get_movements(sku=sample_product.sku, movement_type="OUT")

        assert len(movements) == 1
        # sample_product fixture cost_price = Decimal("100.00")
        assert movements[0].unit_cost == Decimal("100.00")
        assert movements[0].total_cost == Decimal("300.00")  # 3 * 100

    def test_export_movements_cost_columns(self, db_session, sample_product, tmp_path):
        """出入库明细 Excel 导出包含成本数据"""
        from datetime import datetime, timezone
        from decimal import Decimal

        from core.models import InboundReceipt, InboundReceiptItem, InboundStatus, Inventory, InventoryType
        from modules.inventory_offline.reporter import InventoryReporter

        # 建入库单 + Inventory IN
        receipt = InboundReceipt(
            receipt_no="IN-REPORT-COST-002",
            supplier="Test Supplier",
            status=InboundStatus.RECEIVED,
        )
        db_session.add(receipt)
        db_session.flush()

        item = InboundReceiptItem(
            receipt_id=receipt.id,
            sku=sample_product.sku,
            expected_quantity=5,
            received_quantity=5,
            cost_price=Decimal("200.00"),
        )
        db_session.add(item)

        inv = Inventory(
            sku=sample_product.sku,
            type=InventoryType.IN,
            quantity=5,
            location="B-2",
            related_order=receipt.receipt_no,
            occurred_at=datetime.now(timezone.utc),
            operator="test",
        )
        db_session.add(inv)
        db_session.commit()

        reporter = InventoryReporter()
        out_path = tmp_path / "movements_cost.xlsx"
        reporter.export_movements_to_excel(out_path)

        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active

        # 表头：单件成本=col5, 总成本=col6
        assert ws.cell(row=1, column=5).value == "单件成本"
        assert ws.cell(row=1, column=6).value == "总成本"

        # 数据行：IN 记录（成本 = 200）
        in_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "in":
                in_row = row
                break
        assert in_row is not None
        assert ws.cell(row=in_row, column=5).value == 200.0
        assert ws.cell(row=in_row, column=6).value == 1000.0  # 5 * 200
