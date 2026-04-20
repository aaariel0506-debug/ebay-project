"""
tests/test_cost_linker.py

Day 27: CostLinker 测试
"""

import tempfile
from datetime import datetime
from pathlib import Path

from core.models import Transaction, TransactionType
from modules.finance.cost_linker import (
    export_unlinked_xlsx,
    link_costs,
    list_unlinked_orders,
)


class TestLinkCosts:
    """link_costs() 测试"""

    def test_dry_run_does_not_write(self, db_session, sample_product):
        """dry_run=True 时不写入 DB，examined/updated 计数正确"""
        # 造一条 unit_cost=NULL 的 SALE Transaction
        tx = Transaction(
            order_id="ORD-DRY-001",
            sku=sample_product.sku,
            type=TransactionType.SALE,
            amount=150.0,
            currency="USD",
            date=datetime(2026, 4, 10),
            unit_cost=None,  # unlinked
        )
        db_session.add(tx)
        db_session.commit()

        result = link_costs(dry_run=True)

        assert result.examined == 1
        assert result.updated == 1
        assert result.remaining == 0
        assert result.unlinked_skus == []

        # DB 里 unit_cost 仍然是 NULL（dry_run 不写入）
        # 注意：由于 session identity map，直接查询可能命中缓存，需 expire 强制刷新
        db_session.expire_all()
        tx_check = db_session.query(Transaction).filter(
            Transaction.order_id == "ORD-DRY-001"
        ).first()
        assert tx_check.unit_cost is None

    def test_fills_unit_cost(self, db_session, sample_product):
        """非 dry_run 时验证 unit_cost 被填入，total_cost/profit/margin 保持 NULL"""
        tx = Transaction(
            order_id="ORD-FILL-001",
            sku=sample_product.sku,
            type=TransactionType.SALE,
            amount=150.0,
            currency="USD",
            date=datetime(2026, 4, 10),
            unit_cost=None,
        )
        db_session.add(tx)
        db_session.commit()

        result = link_costs(dry_run=False)

        assert result.examined == 1
        assert result.updated == 1
        assert result.remaining == 0

        db_session.expire_all()
        tx_check = db_session.query(Transaction).filter(
            Transaction.order_id == "ORD-FILL-001"
        ).first()
        assert tx_check.unit_cost == float(sample_product.cost_price)
        # total_cost / profit / margin 保持 NULL（link_costs 不填，避免歧义）
        assert tx_check.total_cost is None
        assert tx_check.profit is None
        assert tx_check.margin is None

    def test_unlinked_sku_remaining(self, db_session):
        """SKU 在 Product 表不存在 → remaining=1，unlinked_skus 包含该 SKU"""
        # 造一条 SALE，SKU 不存在于 Product 表
        tx = Transaction(
            order_id="ORD-UNL-001",
            sku="NONEXISTENT-SKU",
            type=TransactionType.SALE,
            amount=200.0,
            currency="USD",
            date=datetime(2026, 4, 10),
            unit_cost=None,
        )
        db_session.add(tx)
        db_session.commit()

        result = link_costs(dry_run=False)

        assert result.examined == 1
        assert result.updated == 0
        assert result.remaining == 1
        assert "NONEXISTENT-SKU" in result.unlinked_skus


class TestListUnlinkedOrders:
    """list_unlinked_orders() 测试"""

    def test_returns_unlinked_sale_transactions(self, db_session, sample_product):
        """unit_cost IS NULL 的 SALE Transaction 被返回"""
        tx = Transaction(
            order_id="ORD-LIST-001",
            sku=sample_product.sku,
            type=TransactionType.SALE,
            amount=180.0,
            currency="USD",
            date=datetime(2026, 4, 12),
            unit_cost=None,
        )
        db_session.add(tx)
        db_session.commit()

        orders = list_unlinked_orders()

        assert len(orders) == 1
        assert orders[0]["order_id"] == "ORD-LIST-001"
        assert orders[0]["sku"] == sample_product.sku
        assert orders[0]["amount"] == 180.0

    def test_excludes_linked_transactions(self, db_session, sample_product):
        """unit_cost 已填的 Transaction 不被返回"""
        tx = Transaction(
            order_id="ORD-LINKED-001",
            sku=sample_product.sku,
            type=TransactionType.SALE,
            amount=180.0,
            currency="USD",
            date=datetime(2026, 4, 12),
            unit_cost=100.0,  # 已关联
        )
        db_session.add(tx)
        db_session.commit()

        orders = list_unlinked_orders()

        assert all(o["order_id"] != "ORD-LINKED-001" for o in orders)


class TestExportUnlinkedXlsx:
    """export_unlinked_xlsx() 测试"""

    def test_exports_xlsx_with_headers(self, db_session, sample_product):
        """生成的 xlsx 包含正确的列名和数据"""
        tx = Transaction(
            order_id="ORD-XLSX-001",
            sku=sample_product.sku,
            type=TransactionType.SALE,
            amount=200.0,
            currency="USD",
            date=datetime(2026, 4, 15),
            unit_cost=None,
        )
        db_session.add(tx)
        db_session.commit()

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "unlinked.xlsx"
            n = export_unlinked_xlsx(path)

            assert n == 1
            assert path.exists()

            # 读取验证
            import openpyxl
            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            assert ws.title == "Unlinked Orders"
            assert ws.cell(1, 1).value == "Order ID"
            assert ws.cell(1, 2).value == "SKU"
            assert ws.cell(1, 3).value == "Amount"
            assert ws.cell(1, 4).value == "Currency"
            assert ws.cell(1, 5).value == "Date"
            assert ws.cell(2, 1).value == "ORD-XLSX-001"
            assert ws.cell(2, 2).value == sample_product.sku

    def test_no_unlinked_returns_zero(self, db_session, sample_product):
        """没有 unlinked Transaction 时返回 0，不报错"""
        # 已关联的 Transaction
        tx = Transaction(
            order_id="ORD-ALL-LINKED-001",
            sku=sample_product.sku,
            type=TransactionType.SALE,
            amount=100.0,
            currency="USD",
            date=datetime(2026, 4, 1),
            unit_cost=100.0,
        )
        db_session.add(tx)
        db_session.commit()

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            n = export_unlinked_xlsx(path)

            assert n == 0
