"""
generator/spreadsheet.py — 生成 Excel 税务报表
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from db.db import fetch_all
from ingest.exchange_rate import batch_get_rates


def generate_report(year: int, output_path: str, month: int | None = None) -> str:
    """
    生成 Excel 税务报表

    Sheet1 列顺序：
    - eBay 订单号 / 成交日期 / 商品名称 / 数量 / 销售价 (USD) / 买家运费 (USD) /
    - eBay 平台费 (USD) / eBay 广告费 (USD) / 快递方式 / 快递单号 /
    - 国际快递费 (USD) / 采购平台 / 采购价 (JPY) / 采购消费税 (JPY) / 采购日期 /
    - 采购订单号 / 汇率 (JPY/USD) / 净利润估算 (USD) / 匹配状态

    净利润 = 销售价 + 买家运费 - eBay 平台费 - eBay 广告费 - 国际快递费 - 采购价 (JPY/汇率)

    Sheet2：未匹配记录（shipments 中 ebay_order_id 为空的行）

    Args:
        year: 报税年份
        output_path: 输出文件路径
        month: 月份 1-12；None 表示全年

    Returns:
        输出文件路径
    """
    wb = Workbook()

    # 创建 Sheet1
    ws1 = wb.active
    ws1.title = "订单明细"

    # 定义表头样式
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = [
        "eBay 订单号", "成交日期", "商品名称", "数量", "销售价 (USD)", "买家运费 (USD)",
        "eBay 平台费 (USD)", "eBay 广告费 (USD)", "快递方式", "快递单号",
        "国际快递费 (USD)", "采购平台", "采购价 (JPY)", "采购消费税 (JPY)", "采购日期",
        "采购订单号", "汇率 (JPY/USD)", "净利润估算 (USD)", "匹配状态"
    ]

    ws1.append(headers)

    # 应用表头样式
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 查询数据：获取指定年份（和可选月份）的 eBay 订单
    if month is not None:
        # 月份模式：过滤年月
        month_padded = f"{month:02d}"
        ebay_orders = fetch_all(
            """
            SELECT * FROM ebay_orders
            WHERE strftime('%Y', sale_date) = ?
              AND strftime('%m', sale_date) = ?
            ORDER BY sale_date
            """,
            (str(year), month_padded)
        )
        report_title = f"eBay 税务报表 {year}年{month_padded}月"
    else:
        # 全年模式
        ebay_orders = fetch_all(
            """
            SELECT * FROM ebay_orders
            WHERE strftime('%Y', sale_date) = ?
            ORDER BY sale_date
            """,
            (str(year),)
        )
        report_title = f"eBay 税务报表 {year}年"
    
    # 无数据时警告
    if not ebay_orders:
        print(f"⚠ {report_title} 无订单数据，生成空报表")

    # 查询所有 shipment
    shipments = fetch_all("SELECT * FROM shipments")
    # 按 ebay_order_id 分组
    shipment_map = {}
    for s in shipments:
        order_id = s.get('ebay_order_id')
        if order_id:
            if order_id not in shipment_map:
                shipment_map[order_id] = []
            shipment_map[order_id].append(s)

    # 查询所有采购
    purchases = fetch_all("SELECT * FROM purchases")

    # 查询匹配关系（包含 FIFO 分配数据）
    links = fetch_all("""
        SELECT purchase_id, ebay_order_id, match_method, confidence, 
               allocated_qty, allocated_cost_jpy, allocated_tax_jpy
        FROM purchase_order_links
    """)
    link_map = {}
    for link in links:
        order_id = link['ebay_order_id']
        if order_id not in link_map:
            link_map[order_id] = []
        link_map[order_id].append(link)

    # 构建采购映射
    purchase_map = {p['id']: p for p in purchases}

    # 批量获取汇率（按订单成交日期）
    order_dates = [order.get('sale_date', '') for order in ebay_orders if order.get('sale_date')]
    rates = batch_get_rates(order_dates)
    fallback_rate = 1.0 / 150.0  # 默认 fallback

    # 生成数据行
    for order in ebay_orders:
        order_id = order['order_id']

        # 获取关联的 shipment
        order_shipments = shipment_map.get(order_id, [])

        # 获取关联的采购
        order_links = link_map.get(order_id, [])
        order_purchases = []
        for link in order_links:
            purchase = purchase_map.get(link['purchase_id'])
            if purchase:
                order_purchases.append({
                    **purchase,
                    'allocated_cost_jpy': link.get('allocated_cost_jpy'),
                    'allocated_tax_jpy': link.get('allocated_tax_jpy'),
                    'allocated_qty': link.get('allocated_qty')
                })

        # 计算汇总值（优先使用 FIFO 分配的成本）
        total_shipping_fee_usd = sum(s.get('shipping_fee_usd') or 0 for s in order_shipments)
        # 如果有 allocated_cost_jpy，使用它；否则使用总采购价
        total_purchase_price_jpy = sum(
            p.get('allocated_cost_jpy') if p.get('allocated_cost_jpy') is not None else p.get('total_price_jpy') or 0
            for p in order_purchases
        )
        total_tax_jpy = sum(
            p.get('allocated_tax_jpy') if p.get('allocated_tax_jpy') is not None else p.get('tax_jpy') or 0
            for p in order_purchases
        )

        # 获取第一个 shipment 的信息
        carrier = order_shipments[0]['carrier'] if order_shipments else ''
        tracking_number = order_shipments[0]['tracking_number'] if order_shipments else ''

        # 获取第一个采购的信息
        purchase_platform = order_purchases[0]['platform'] if order_purchases else ''
        purchase_date = order_purchases[0]['purchase_date'] if order_purchases else ''
        purchase_order_number = order_purchases[0]['order_number'] if order_purchases else ''

        # 获取该订单的汇率
        order_date = order.get('sale_date', '')
        rate = rates.get(order_date, fallback_rate) if order_date else fallback_rate
        # API 返回 1 JPY = X USD，但我们需要 1 USD = Y JPY 来显示
        rate_display = 1.0 / rate if rate > 0 else 150.0

        # 计算净利润
        sale_price = order.get('sale_price_usd') or 0
        buyer_shipping = order.get('shipping_charged_usd') or 0
        ebay_fee = order.get('ebay_fee_usd') or 0
        ebay_ad_fee = order.get('ebay_ad_fee_usd') or 0

        net_profit = (
            sale_price
            + buyer_shipping
            - ebay_fee
            - ebay_ad_fee
            - total_shipping_fee_usd
            - (total_purchase_price_jpy * rate)
        )

        # 匹配状态
        if order_purchases and order_shipments:
            match_status = "已匹配"
        elif order_purchases:
            match_status = "部分匹配（缺快递）"
        elif order_shipments:
            match_status = "部分匹配（缺采购）"
        else:
            match_status = "未匹配"

        row = [
            order_id,
            order.get('sale_date') or '',
            order.get('item_title') or '',
            order.get('quantity') or 1,
            sale_price,
            buyer_shipping,
            ebay_fee,
            ebay_ad_fee,
            carrier,
            tracking_number,
            total_shipping_fee_usd,
            purchase_platform,
            total_purchase_price_jpy,
            total_tax_jpy,
            purchase_date,
            purchase_order_number,
            round(rate_display, 2),  # 汇率列：1 USD = X JPY
            round(net_profit, 2),
            match_status,
        ]

        ws1.append(row)

    # 创建 Sheet2：库存表
    ws2 = wb.create_sheet(title="库存")
    inventory_headers = [
        "SKU", "商品名 (日)", "商品名 (英)", "总采购数量", "已销售数量", 
        "剩余数量", "总成本 (JPY)", "总税额 (JPY)", "平均单位成本 (JPY)"
    ]
    ws2.append(inventory_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 查询库存数据
    inventory = fetch_all("SELECT * FROM inventory ORDER BY item_sku")
    for item in inventory:
        ws2.append([
            item.get('item_sku') or '',
            item.get('item_name') or '',
            item.get('item_name_en') or '',
            item.get('total_quantity') or 0,
            item.get('sold_quantity') or 0,
            item.get('remaining_quantity') or 0,
            item.get('total_cost_jpy') or 0,
            item.get('total_tax_jpy') or 0,
            round(item.get('average_cost_per_unit') or 0, 2)
        ])
    
    # 创建 Sheet3：未匹配记录
    ws3 = wb.create_sheet(title="未匹配记录")

    # 未匹配 shipment 表头
    shipment_headers = [
        "ID", "carrier", "tracking_number", "ebay_order_id",
        "ship_date", "shipping_fee_usd", "cpass_transaction_id"
    ]
    ws3.append(shipment_headers)

    # 应用表头样式
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 查询未匹配的 shipment（ebay_order_id 为空）
    unmatched_shipments = fetch_all(
        "SELECT * FROM shipments WHERE ebay_order_id IS NULL ORDER BY ship_date"
    )

    for shipment in unmatched_shipments:
        row = [
            shipment.get('id') or '',
            shipment.get('carrier') or '',
            shipment.get('tracking_number') or '',
            shipment.get('ebay_order_id') or '',
            shipment.get('ship_date') or '',
            shipment.get('shipping_fee_usd') or '',
            shipment.get('cpass_transaction_id') or '',
        ]
        ws3.append(row)

    # 保存文件
    wb.save(output_path)

    return output_path
