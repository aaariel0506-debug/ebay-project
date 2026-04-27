from datetime import datetime
from decimal import Decimal

import pytest
from core.models import Order, OrderStatus, Transaction, TransactionType
from modules.finance.cpass_importer import (
    CpassImportError,
    _import_with_session,
    _parse_amount,
)
from openpyxl import Workbook


def _make_wb(path, order_rows=None, fee_rows=None, order_headers=None, fee_headers=None, include_order=True, include_fee=True):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    if include_order:
        ws = wb.create_sheet("Order")
        headers = order_headers or ["Order No.", "Tracking No.", "Payable Amount"]
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=h)
        for r_idx, row in enumerate(order_rows or [], start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    if include_fee:
        ws = wb.create_sheet("Fee Details")
        headers = fee_headers or ["Tracking No.", "Amount (JPY)"]
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=h)
        for r_idx, row in enumerate(fee_rows or [], start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.create_sheet("SKU Details")
    wb.save(path)


def _order(db_session, oid, tracking_no=None, status=OrderStatus.SHIPPED, created_at=None):
    obj = Order(
        ebay_order_id=oid,
        sale_price=Decimal("0"),
        shipping_cost=Decimal("0"),
        ebay_fee=Decimal("0"),
        buyer_country="US",
        status=status,
        order_date=datetime(2026, 4, 15, 10),
        tracking_no=tracking_no,
    )
    if created_at is not None:
        obj.created_at = created_at
        obj.updated_at = created_at
    db_session.add(obj)
    db_session.flush()
    return obj


class TestParseAmount:
    def test_parse_string_with_comma(self):
        assert _parse_amount("4,461") == Decimal("4461")

    def test_parse_signed_string(self):
        assert _parse_amount("+1,367") == Decimal("1367")
        assert _parse_amount("-4,461") == Decimal("-4461")

    def test_parse_none_returns_zero(self):
        assert _parse_amount(None) == Decimal("0")

    def test_parse_int_or_float(self):
        assert _parse_amount(4461) == Decimal("4461")
        assert _parse_amount(4461.5) == Decimal("4461.5")

    def test_parse_garbage_returns_zero(self):
        assert _parse_amount("abc") == Decimal("0")


class TestStructureValidation:
    def test_missing_order_sheet_raises(self, tmp_path, db_session):
        path = tmp_path / "x.xlsx"
        _make_wb(path, include_order=False)
        with pytest.raises(CpassImportError):
            _import_with_session(db_session, str(path))

    def test_missing_fee_details_sheet_raises(self, tmp_path, db_session):
        path = tmp_path / "x.xlsx"
        _make_wb(path, include_fee=False)
        with pytest.raises(CpassImportError):
            _import_with_session(db_session, str(path))

    def test_order_sheet_missing_order_no_col_raises(self, tmp_path, db_session):
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_headers=["Wrong Col"], fee_rows=[["T1", "100"]])
        with pytest.raises(CpassImportError):
            _import_with_session(db_session, str(path))


class TestMatching:
    def test_matched_tracking_writes_shipping_actual(self, tmp_path, db_session):
        _order(db_session, "O1", tracking_no="CPASS-1")
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_rows=[["CPASS-1", "T", "0"]], fee_rows=[["CPASS-1", "4,461"], ["CPASS-1", "+1,367"]])

        result = _import_with_session(db_session, str(path))
        tx = db_session.query(Transaction).filter(Transaction.order_id == "O1", Transaction.type == TransactionType.SHIPPING_ACTUAL).one()
        assert result.total_rows == 1
        assert result.matched == 1
        assert result.written == 1
        assert Decimal(str(tx.amount_jpy)) == Decimal("5828")
        assert Decimal(str(tx.amount)) == Decimal("0")

    def test_unmatched_no_tracking_increments_counter(self, tmp_path, db_session):
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_rows=[["MISS-1", "T", "0"]], fee_rows=[["MISS-1", "100"]])
        result = _import_with_session(db_session, str(path))
        assert result.unmatched_no_tracking == 1
        assert db_session.query(Transaction).count() == 0

    def test_cancelled_order_skipped_with_warning(self, tmp_path, db_session):
        _order(db_session, "O1", tracking_no="CPASS-1", status=OrderStatus.CANCELLED)
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_rows=[["CPASS-1", "T", "0"]], fee_rows=[["CPASS-1", "100"]])
        result = _import_with_session(db_session, str(path))
        assert result.unmatched_cancelled == 1
        assert db_session.query(Transaction).count() == 0

    def test_zero_payable_still_writes_amount_zero(self, tmp_path, db_session):
        _order(db_session, "O1", tracking_no="CPASS-1")
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_rows=[["CPASS-1", "T", 0]], fee_rows=[["CPASS-1", 0]])
        result = _import_with_session(db_session, str(path))
        tx = db_session.query(Transaction).filter_by(order_id="O1", type=TransactionType.SHIPPING_ACTUAL).one()
        assert result.skipped_zero == 1
        assert Decimal(str(tx.amount_jpy)) == Decimal("0")

    def test_multi_order_with_same_tracking_takes_latest(self, tmp_path, db_session):
        _order(db_session, "OLD", tracking_no="CPASS-1", created_at=datetime(2026, 4, 15, 10))
        _order(db_session, "NEW", tracking_no="CPASS-1", created_at=datetime(2026, 4, 15, 11))
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_rows=[["CPASS-1", "T", 0]], fee_rows=[["CPASS-1", 200]])
        _import_with_session(db_session, str(path))
        assert db_session.query(Transaction).filter_by(order_id="NEW", type=TransactionType.SHIPPING_ACTUAL).count() == 1
        assert db_session.query(Transaction).filter_by(order_id="OLD", type=TransactionType.SHIPPING_ACTUAL).count() == 0


class TestIdempotency:
    def test_rerun_overwrites_old_shipping_actual(self, tmp_path, db_session):
        _order(db_session, "O1", tracking_no="CPASS-1")
        path1 = tmp_path / "x1.xlsx"
        path2 = tmp_path / "x2.xlsx"
        _make_wb(path1, order_rows=[["CPASS-1", "T", 0]], fee_rows=[["CPASS-1", 100]])
        _make_wb(path2, order_rows=[["CPASS-1", "T", 0]], fee_rows=[["CPASS-1", 250]])
        _import_with_session(db_session, str(path1))
        _import_with_session(db_session, str(path2))
        rows = db_session.query(Transaction).filter_by(order_id="O1", type=TransactionType.SHIPPING_ACTUAL).all()
        assert len(rows) == 1
        assert Decimal(str(rows[0].amount_jpy)) == Decimal("250")


class TestDryRun:
    def test_dry_run_does_not_write_db(self, tmp_path, db_session):
        _order(db_session, "O1", tracking_no="CPASS-1")
        path = tmp_path / "x.xlsx"
        _make_wb(path, order_rows=[["CPASS-1", "T", 0]], fee_rows=[["CPASS-1", 100]])
        result = _import_with_session(db_session, str(path), dry_run=True)
        assert result.matched == 1
        assert result.written == 0
        assert db_session.query(Transaction).count() == 0
