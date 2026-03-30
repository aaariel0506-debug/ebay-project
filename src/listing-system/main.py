#!/usr/bin/env python3
"""
eBay Listing 自动化 — 主入口
读取 Excel 商品表 → 校验 → 创建 Listing → 写回结果

用法:
    python3 main.py                          # 默认读取 商品数据.xlsx
    python3 main.py products.xlsx            # 指定 Excel 文件
    python3 main.py products.xlsx --publish  # 创建后直接发布（谨慎使用）
    python3 main.py --test                   # 测试 API 连接
"""

import sys
import logging
from pathlib import Path
from datetime import datetime

# ─── 日志配置 ──────────────────────────────────────────
LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)

log_file = LOG_DIR / f"listing_{datetime.now():%Y%m%d_%H%M%S}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("main")

# ─── 导入模块 ──────────────────────────────────────────
from ebay_client import EbayClient
from data_validator import validate_batch
from listing_creator import ListingCreator

try:
    import openpyxl
except ImportError:
    logger.error("需要安装 openpyxl: pip install openpyxl --break-system-packages")
    sys.exit(1)


def read_excel(file_path: str) -> list:
    """读取 Excel 商品数据"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())

    # 读取数据行
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        for i, val in enumerate(row):
            if i < len(headers):
                item[headers[i]] = val
        # 跳过已成功处理的行
        status = str(item.get("status", "")).strip().upper()
        if status == "SUCCESS":
            continue
        # 跳过完全空白的行
        if not item.get("sku"):
            continue
        rows.append(item)

    logger.info(f"从 {file_path} 读取到 {len(rows)} 条待处理商品")
    return rows


def write_results(file_path: str, results: dict):
    """将结果写回 Excel（更新 status, offer_id, error_msg, processed_at 列）"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 找到表头列索引
    header_map = {}
    for i, cell in enumerate(ws[1], 1):
        header_map[str(cell.value or "").strip().lower()] = i

    # 确保结果列存在
    result_cols = ["status", "offer_id", "listing_id", "error_msg", "processed_at"]
    max_col = max(header_map.values()) if header_map else 0
    for col_name in result_cols:
        if col_name not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=col_name)
            header_map[col_name] = max_col

    # 写入结果
    sku_col = header_map.get("sku", 1)
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        sku_val = str(row[sku_col - 1].value or "").strip()
        if sku_val in results:
            r = results[sku_val]
            ws.cell(row=row[0].row, column=header_map["status"], value=r.status_text)
            ws.cell(row=row[0].row, column=header_map["offer_id"], value=r.offer_id)
            ws.cell(row=row[0].row, column=header_map["listing_id"], value=r.listing_id)
            ws.cell(row=row[0].row, column=header_map["error_msg"], value=r.error)
            ws.cell(
                row=row[0].row,
                column=header_map["processed_at"],
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )

    wb.save(file_path)
    logger.info(f"结果已写回 {file_path}")


def main():
    logger.info("=" * 60)
    logger.info("eBay Listing 自动化启动")
    logger.info("=" * 60)

    # 解析参数
    args = sys.argv[1:]
    publish = "--publish" in args
    test_mode = "--test" in args
    excel_file = None

    for arg in args:
        if arg.endswith(".xlsx") or arg.endswith(".xls"):
            excel_file = arg

    if not excel_file:
        excel_file = str(Path(__file__).parent / "商品数据.xlsx")

    # 初始化客户端
    client = EbayClient()
    logger.info(f"环境: {client.config.get('environment', 'sandbox')}")
    logger.info(f"API:  {client.api_base}")
    # 自动发布：必须同时满足：命令行--publish、配置auto_publish=true、配置require_review=false
    config_auto_publish = client.config.get('listing_defaults', {}).get('auto_publish', False)
    require_review = client.config.get('workflow', {}).get('require_review', True)
    auto_publish = publish and config_auto_publish and not require_review
    
    if publish and not auto_publish:
        if require_review:
            logger.warning("配置 workflow.require_review=true，强制进入预审核模式")
        elif not config_auto_publish:
            logger.warning("配置 listing_defaults.auto_publish=false，强制进入预审核模式")
    
    logger.info(f"自动发布：{'是' if auto_publish else '否（草稿模式，需预审核）'}")

    # 测试模式
    if test_mode:
        logger.info("\n--- API 连接测试 ---")
        if client.test_connection():
            logger.info("连接正常，可以开始使用")
        else:
            logger.error("连接失败，请检查配置")
        return

    # 读取 Excel
    if not Path(excel_file).exists():
        logger.error(f"Excel 文件不存在: {excel_file}")
        logger.info(f"请将商品数据放入: {excel_file}")
        logger.info("或指定文件路径: python3 main.py your_file.xlsx")
        return

    rows = read_excel(excel_file)
    if not rows:
        logger.info("没有待处理的商品（可能全部已成功）")
        return

    # 数据校验
    logger.info(f"\n--- 数据校验 ({len(rows)} 条) ---")
    validations = validate_batch(rows)

    valid_rows = []
    for row, v in zip(rows, validations):
        if v.valid:
            valid_rows.append(row)
        else:
            logger.warning(f"跳过 SKU {v.sku}: {'; '.join(v.errors)}")

    if not valid_rows:
        logger.error("没有通过校验的商品，请修正 Excel 后重试")
        return

    logger.info(f"{len(valid_rows)}/{len(rows)} 条通过校验，开始创建 Listing")

    # 创建 Listing
    creator = ListingCreator(client)
    all_results = {}

    for i, item in enumerate(valid_rows, 1):
        sku = str(item["sku"]).strip()
        logger.info(f"\n--- [{i}/{len(valid_rows)}] 处理 SKU: {sku} ---")

        result = creator.create_listing(item, auto_publish=auto_publish)
        all_results[sku] = result

        if result.success:
            logger.info(f"[{sku}] 成功! offer_id={result.offer_id}")
        else:
            logger.error(f"[{sku}] 失败: {result.error}")

    # 写回结果
    write_results(excel_file, all_results)

    # 汇总
    success_count = sum(1 for r in all_results.values() if r.success)
    fail_count = len(all_results) - success_count

    logger.info("\n" + "=" * 60)
    logger.info(f"处理完成! 成功: {success_count}, 失败: {fail_count}")
    logger.info(f"日志文件: {log_file}")

    if not publish and success_count > 0:
        logger.info(
            f"\n所有成功的 Listing 为草稿状态 (UNPUBLISHED)")
        logger.info(f"请在 Seller Hub → Drafts 中查看并手动发布:")
        env = client.config.get("environment", "sandbox")
        if env == "production":
            logger.info("  https://www.ebay.com/sh/lst/drafts")
        else:
            logger.info("  https://www.sandbox.ebay.com/sh/lst/drafts")

    logger.info("=" * 60)


if __name__ == "__main__":
    main()
