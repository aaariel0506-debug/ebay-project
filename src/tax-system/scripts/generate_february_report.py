#!/usr/bin/env python3
"""
生成 2026 年 2 月份 eBay 报税报表
"""
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

DB_PATH = '/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/orders.db'
OUTPUT_PATH = f'/Users/arielhe/.openclaw/workspace/ebay-tax-system/data/outputs/2026-02/tax_report_2026-02_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

def generate_february_report():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 创建工作簿
    wb = Workbook()
    
    # === Sheet 1: 财务摘要 ===
    ws_summary = wb.active
    ws_summary.title = "财务摘要"
    
    # 表头样式
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # 写入标题
    ws_summary['A1'] = "2026 年 2 月 eBay 报税统计"
    ws_summary['A1'].font = Font(bold=True, size=16)
    
    # 查询统计数据
    cursor.execute("SELECT COUNT(*) as cnt FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'")
    order_count = cursor.fetchone()['cnt'] or 0
    
    cursor.execute("SELECT SUM(sale_price_usd) as total FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'")
    total_sales = cursor.fetchone()['total'] or 0
    
    cursor.execute("SELECT SUM(shipping_charged_usd) as total FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02'")
    total_shipping = cursor.fetchone()['total'] or 0
    
    cursor.execute("""
        SELECT SUM(shipping_fee_usd) as total FROM shipments 
        WHERE ebay_order_id IN (SELECT order_id FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02')
    """)
    total_shipping_cost = cursor.fetchone()['total'] or 0
    
    cursor.execute("""
        SELECT SUM(allocated_cost_jpy) as total FROM purchase_order_links 
        WHERE ebay_order_id IN (SELECT order_id FROM ebay_orders WHERE strftime('%Y-%m', sale_date) = '2026-02')
    """)
    total_purchase_cost_jpy = cursor.fetchone()['total'] or 0
    
    # 估算汇率
    exchange_rate = 150.0
    total_purchase_cost_usd = total_purchase_cost_jpy / exchange_rate
    
    # 计算净利润
    net_profit = total_sales + total_shipping - total_shipping_cost - total_purchase_cost_usd
    
    # 写入统计数据
    summary_data = [
        ["项目", "数值"],
        ["订单数量", order_count],
        ["销售总额 (USD)", f"${total_sales:.2f}"],
        ["买家运费收入 (USD)", f"${total_shipping:.2f}"],
        ["快递成本 (USD)", f"${total_shipping_cost:.2f}"],
        ["采购成本 (JPY)", f"¥{total_purchase_cost_jpy:,.0f}"],
        ["采购成本 (USD)", f"${total_purchase_cost_usd:.2f}"],
        ["估算汇率", f"{exchange_rate} JPY/USD"],
        ["", ""],
        ["净利润估算 (USD)", f"${net_profit:.2f}"],
    ]
    
    for i, row in enumerate(summary_data, start=3):
        ws_summary.append(row)
    
    # === Sheet 2: 订单明细 ===
    ws_orders = wb.create_sheet(title="订单明细")
    
    order_headers = [
        "eBay 订单号", "成交日期", "商品名称", "数量", "销售价 (USD)", 
        "买家运费 (USD)", "eBay 平台费 (USD)", "eBay 广告费 (USD)",
        "快递方式", "快递单号", "国际快递费 (USD)",
        "采购平台", "采购价 (JPY)", "采购日期", "采购订单号",
        "汇率 (JPY/USD)", "净利润估算 (USD)", "匹配状态"
    ]
    ws_orders.append(order_headers)
    
    # 应用表头样式
    for cell in ws_orders[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 查询 2 月份订单
    cursor.execute("""
        SELECT * FROM ebay_orders 
        WHERE strftime('%Y-%m', sale_date) = '2026-02'
        ORDER BY sale_date
    """)
    orders = cursor.fetchall()
    
    for order in orders:
        order_id = order['order_id']
        
        # 获取关联的 shipment
        cursor.execute("SELECT * FROM shipments WHERE ebay_order_id = ?", (order_id,))
        shipment = cursor.fetchone()
        
        # 获取关联的采购
        cursor.execute("""
            SELECT pol.*, p.platform, p.total_price_jpy, p.order_number, p.purchase_date
            FROM purchase_order_links pol
            JOIN purchases p ON pol.purchase_id = p.id
            WHERE pol.ebay_order_id = ?
        """, (order_id,))
        purchases = cursor.fetchall()
        
        # 计算汇总值
        shipping_fee_usd = shipment['shipping_fee_usd'] if shipment else 0
        total_purchase_jpy = sum(p['allocated_cost_jpy'] or p['total_price_jpy'] or 0 for p in purchases)
        
        # 匹配状态
        if purchases and shipment:
            match_status = "已匹配"
        elif purchases:
            match_status = "部分匹配（缺快递）"
        elif shipment:
            match_status = "部分匹配（缺采购）"
        else:
            match_status = "未匹配"
        
        # 计算净利润
        sale_price = order['sale_price_usd'] or 0
        buyer_shipping = order['shipping_charged_usd'] or 0
        ebay_fee = order['ebay_fee_usd'] or 0
        ebay_ad_fee = order['ebay_ad_fee_usd'] or 0
        
        net_profit = sale_price + buyer_shipping - ebay_fee - ebay_ad_fee - shipping_fee_usd - (total_purchase_jpy / exchange_rate)
        
        row = [
            order_id,
            order['sale_date'] or '',
            order['item_title'] or '',
            order['quantity'] or 1,
            sale_price,
            buyer_shipping,
            ebay_fee,
            ebay_ad_fee,
            shipment['carrier'] if shipment else '',
            shipment['tracking_number'] if shipment else '',
            shipping_fee_usd,
            purchases[0]['platform'] if purchases else '',
            total_purchase_jpy,
            purchases[0]['purchase_date'] if purchases else '',
            purchases[0]['order_number'] if purchases else '',
            exchange_rate,
            round(net_profit, 2),
            match_status
        ]
        ws_orders.append(row)
    
    # === Sheet 3: 快递记录 ===
    ws_shipments = wb.create_sheet(title="快递记录")
    
    shipment_headers = [
        "ID", "追踪号", "关联订单", "发货日期", "快递方式", 
        "运费 (USD)", "匹配状态", "确认人"
    ]
    ws_shipments.append(shipment_headers)
    
    for cell in ws_shipments[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 查询 2 月份相关的快递
    cursor.execute("""
        SELECT s.* FROM shipments s
        WHERE s.ebay_order_id IN (
            SELECT order_id FROM ebay_orders 
            WHERE strftime('%Y-%m', sale_date) = '2026-02'
        )
        ORDER BY s.ship_date
    """)
    shipments = cursor.fetchall()
    
    for s in shipments:
        ws_shipments.append([
            s['id'],
            s['tracking_number'],
            s['ebay_order_id'],
            s['ship_date'],
            s['carrier'],
            s['shipping_fee_usd'],
            s['match_method'] or '待匹配',
            s['confirmed_by'] or ''
        ])
    
    # === Sheet 4: 采购记录 ===
    ws_purchases = wb.create_sheet(title="采购记录")
    
    purchase_headers = [
        "采购 ID", "平台", "采购日期", "商品名", "ASIN",
        "数量", "单价 (JPY)", "总价 (JPY)", "税额 (JPY)",
        "订单号", "匹配状态"
    ]
    ws_purchases.append(purchase_headers)
    
    for cell in ws_purchases[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 查询已匹配的采购
    cursor.execute("""
        SELECT p.*, pol.ebay_order_id, pol.allocated_cost_jpy
        FROM purchases p
        JOIN purchase_order_links pol ON p.id = pol.purchase_id
        WHERE pol.ebay_order_id IN (
            SELECT order_id FROM ebay_orders 
            WHERE strftime('%Y-%m', sale_date) = '2026-02'
        )
        ORDER BY p.purchase_date
    """)
    purchases = cursor.fetchall()
    
    for p in purchases:
        ws_purchases.append([
            p['id'],
            p['platform'],
            p['purchase_date'],
            p['item_name'],
            p['item_sku'],
            p['quantity'],
            p['unit_price_jpy'],
            p['total_price_jpy'],
            p['tax_jpy'],
            p['order_number'],
            '已匹配' if p['ebay_order_id'] else '未匹配'
        ])
    
    # 保存文件
    import os
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    
    conn.close()
    
    print(f"✅ 2 月份报税报表已生成：{OUTPUT_PATH}")
    print(f"📊 订单数量：{order_count}")
    print(f"💰 销售总额：${total_sales:.2f}")
    print(f"📦 采购成本：¥{total_purchase_cost_jpy:,.0f} (约 ${total_purchase_cost_usd:.2f})")
    print(f"📈 净利润估算：${net_profit:.2f}")


if __name__ == '__main__':
    generate_february_report()
