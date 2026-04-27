"""
cpass 运费导入器 — Day 31.5-B

把 cpass 后台导出的 "cpass运单费用明细.xlsx" 解析为 Transaction.SHIPPING_ACTUAL 记录。

业务规则见 docs/finance-semantics.md §3.shipping_actual。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from core.database.connection import get_session
from core.models import Order, Transaction
from core.models.order import OrderStatus
from core.models.transaction import TransactionType
from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session


class CpassImportError(Exception):
    """cpass xlsx 结构异常 — 缺 sheet / header 错位等"""


@dataclass
class CpassImportResult:
    total_rows: int = 0
    matched: int = 0
    unmatched_no_tracking: int = 0
    unmatched_cancelled: int = 0
    written: int = 0
    skipped_zero: int = 0
    errors: list[str] = field(default_factory=list)


def _parse_amount(raw: object) -> Decimal:
    """把 '4,461' / '+1,367' / '-4,461' / 4461.0 / None 解析为 Decimal,失败返回 Decimal('0')。"""
    if raw is None:
        return Decimal("0")
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    s = (
        str(raw)
        .strip()
        .replace(",", "")
        .replace("+", "")
        .replace("¥", "")
        .replace("$", "")
        .replace("JPY", "")
    )
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _build_header_index(ws) -> dict[str, int]:
    """读 ws 第 1 行,返回 {header_name: 1-based col idx}。"""
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            headers[str(v).strip()] = c
    return headers


def _sum_fee_details_for_tracking(ws_fees, tracking_no: str, hdr: dict[str, int]) -> Decimal:
    """在 Fee Details sheet 里找出所有 Tracking No.= tracking_no 的行,Σ Amount (JPY)。"""
    tracking_col = hdr.get("Tracking No.")
    amount_col = hdr.get("Amount (JPY)")
    if not tracking_col or not amount_col:
        raise CpassImportError(
            f"Fee Details sheet missing required columns. headers={list(hdr.keys())}"
        )
    total = Decimal("0")
    for r in range(2, ws_fees.max_row + 1):
        row_tracking = ws_fees.cell(row=r, column=tracking_col).value
        if row_tracking and str(row_tracking).strip() == tracking_no:
            total += _parse_amount(ws_fees.cell(row=r, column=amount_col).value)
    return total


def _import_with_session(sess: Session, file_path: str, dry_run: bool = False) -> CpassImportResult:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"cpass xlsx not found: {file_path}")

    wb = load_workbook(str(path), data_only=True, read_only=False)
    if "Order" not in wb.sheetnames:
        raise CpassImportError(f"missing 'Order' sheet. found: {wb.sheetnames}")
    if "Fee Details" not in wb.sheetnames:
        raise CpassImportError(f"missing 'Fee Details' sheet. found: {wb.sheetnames}")

    ws_order = wb["Order"]
    ws_fees = wb["Fee Details"]
    order_hdr = _build_header_index(ws_order)
    fees_hdr = _build_header_index(ws_fees)

    for col in ["Order No."]:
        if col not in order_hdr:
            raise CpassImportError(
                f"Order sheet missing required col {col!r}. headers={list(order_hdr.keys())}"
            )

    result = CpassImportResult()
    now = datetime.utcnow()

    for r in range(2, ws_order.max_row + 1):
        cpass_order_no = ws_order.cell(row=r, column=order_hdr["Order No."]).value
        if not cpass_order_no:
            continue
        cpass_order_no = str(cpass_order_no).strip()
        result.total_rows += 1

        stmt = (
            select(Order)
            .where(Order.tracking_no == cpass_order_no)
            .order_by(Order.created_at.desc())
        )
        orders = list(sess.execute(stmt).scalars())

        if not orders:
            result.unmatched_no_tracking += 1
            logger.warning(
                "[cpass] tracking={} 在 Order 表里未找到匹配的 ebay_order_id",
                cpass_order_no,
            )
            continue

        if len(orders) > 1:
            logger.warning(
                "[cpass] tracking={} 匹配到多笔 ebay_order: {},取最近创建那笔",
                cpass_order_no,
                [o.ebay_order_id for o in orders],
            )

        ebay_order = orders[0]
        if ebay_order.status == OrderStatus.CANCELLED:
            result.unmatched_cancelled += 1
            logger.warning(
                "[cpass] tracking={} 对应订单 {} 已 CANCELLED,跳过",
                cpass_order_no,
                ebay_order.ebay_order_id,
            )
            continue

        try:
            amount_jpy = _sum_fee_details_for_tracking(ws_fees, cpass_order_no, fees_hdr)
        except CpassImportError:
            raise
        except Exception as e:
            result.errors.append(f"row {r} tracking={cpass_order_no} fee 解析失败: {e}")
            continue

        result.matched += 1
        if amount_jpy == Decimal("0"):
            result.skipped_zero += 1

        if dry_run:
            continue

        sess.query(Transaction).filter(
            Transaction.order_id == ebay_order.ebay_order_id,
            Transaction.type == TransactionType.SHIPPING_ACTUAL,
        ).delete(synchronize_session=False)

        sess.add(
            Transaction(
                order_id=ebay_order.ebay_order_id,
                sku=None,
                type=TransactionType.SHIPPING_ACTUAL,
                amount=Decimal("0"),
                currency="JPY",
                amount_jpy=amount_jpy,
                date=now,
                note=f"cpass tracking={cpass_order_no}",
            )
        )
        result.written += 1

    if not dry_run:
        sess.flush()
    return result


def import_cpass_shipping(file_path: str, dry_run: bool = False) -> CpassImportResult:
    """
    主入口 — 解析 cpass xlsx 并写入 SHIPPING_ACTUAL Transaction。
    """
    with get_session() as sess:
        return _import_with_session(sess, file_path, dry_run=dry_run)
